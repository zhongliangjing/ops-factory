"""
XLSX Builder for Comprehensive Quality Report.
Creates a 10-sheet Excel workbook with data tables, native Excel charts,
and AI-powered insights per sheet.
"""

import io
import os
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from config import OUTPUT_DIR
from analyzer import ComprehensiveResult
from xlsx_theme import (
    XlsxStyles, ROW_HEIGHTS, COL_WIDTHS,
    format_duration, format_pct, format_number,
    sla_level, rating_text, efficiency_level,
    CHART_ROWS_STANDARD, CHART_ROWS_SMALL, TAB_COLORS,
)
from xlsx_analyzer import (
    XlsxDetailAnalyzer, ActionPlanRow,
)
from xlsx_chart_native import NativeChartEngine
from xlsx_chart_matplotlib import MatplotlibChartEngine

logger = logging.getLogger(__name__)

# ── Sheet names ──────────────────────────────────────────────────────────────

SHEET_NAMES = {
    "zh": [
        "执行摘要", "INC_事件分析", "INC_SLA分析", "CHG_变更分析", "SRQ_请求分析",
        "PRO_问题分析", "CRO_跨流程关联", "CRO_人员与效率", "CRO_时间维度", "行动计划",
    ],
    "en": [
        "Executive Summary", "INC_Analysis", "INC_SLA",
        "CHG_Analysis", "SRQ_Analysis", "PRO_Analysis",
        "CRO_Cross-Process", "CRO_Personnel", "CRO_Time Analysis",
        "Action Plan",
    ],
}


class XlsxBuilder:
    """Assembles a 10-sheet Excel workbook from analysis results."""

    def __init__(
        self,
        result: ComprehensiveResult,
        incidents_df,
        changes_df,
        requests_df,
        problems_df,
        sla_map: Dict,
        insights: Dict[str, str],
        language: str = "en",
        chart_engine: str = "native",
    ):
        self.result = result
        self.incidents_df = incidents_df
        self.changes_df = changes_df
        self.requests_df = requests_df
        self.problems_df = problems_df
        self.sla_map = sla_map or {}
        self.insights = insights or {}
        self.language = language

        self.styles = XlsxStyles(language)
        self.detail = XlsxDetailAnalyzer(
            incidents_df, changes_df, requests_df, problems_df,
            sla_map, result, language,
        )
        self.wb: Optional[openpyxl.Workbook] = None
        self.charts = None  # initialized in build() after wb is created
        self.chart_engine_type = chart_engine  # "native" or "matplotlib"

    # ─── Utility methods ─────────────────────────────────────────────────

    def _write_title(self, ws, title: str, subtitle: str = "", row: int = 1) -> int:
        """Write sheet title and optional subtitle. Returns next available row."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.styles.font_h1
        cell.alignment = self.styles.align_h1
        ws.row_dimensions[row].height = ROW_HEIGHTS["h1"]
        row += 1

        if not subtitle:
            subtitle = f"{self.result.start_date} — {self.result.end_date}"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=subtitle)
        cell.font = self.styles.font_h2
        cell.alignment = self.styles.align_left
        ws.row_dimensions[row].height = ROW_HEIGHTS["h2"]
        row += 1

        # spacer
        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        return row + 1

    def _write_section(self, ws, title: str, row: int) -> int:
        """Write a section heading. Returns next row."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.styles.font_h3
        cell.alignment = self.styles.align_left
        ws.row_dimensions[row].height = ROW_HEIGHTS["h3"]
        return row + 1

    def _write_table(self, ws, headers: List[str], data: List[List], row: int,
                     col_widths: Optional[List[float]] = None) -> int:
        """Write a table with header and data rows. Returns next row after table."""
        # Header row
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = self.styles.font_th
            cell.fill = self.styles.fill_header
            cell.alignment = self.styles.align_center
            cell.border = self.styles.border_header
        ws.row_dimensions[row].height = ROW_HEIGHTS["th"]
        row += 1

        # Data rows with zebra striping
        for ri, data_row in enumerate(data):
            fill = self.styles.fill_zebra if ri % 2 == 0 else self.styles.fill_white
            for ci, val in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = self.styles.font_td
                cell.fill = fill
                cell.border = self.styles.border_row
                # Right-align numbers
                if isinstance(val, (int, float)):
                    cell.alignment = self.styles.align_right
                    cell.font = self.styles.font_td_num
                else:
                    cell.alignment = self.styles.align_left
            ws.row_dimensions[row].height = ROW_HEIGHTS["td"]
            row += 1

        # Apply column widths
        if col_widths:
            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
        else:
            # Auto-width: measure header + data content
            all_rows = [headers] + data
            for ci in range(1, len(headers) + 1):
                max_len = max(
                    (len(str(r[ci - 1])) if ci - 1 < len(r) else 0 for r in all_rows),
                    default=8,
                )
                ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 4, 10), 50)

        # spacer
        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        return row + 1

    def _write_chart_desc(self, ws, text: str, row: int) -> int:
        """Write a brief one-line description before a chart."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name="Calibri", size=9, italic=True, color="64748b")
        cell.alignment = self.styles.align_left
        ws.row_dimensions[row].height = 16
        return row + 1

    def _add_chart(self, ws, chart_fn, *args, row=None, small=False, **kwargs):
        """Call a chart engine method, add result to worksheet. Returns next row.

        Handles both native openpyxl chart objects and matplotlib PNG BytesIO.
        For matplotlib PNGs, scales proportionally based on target height to
        prevent overlapping while preserving aspect ratio.
        """
        if row is None:
            raise ValueError("row must be specified")
        try:
            result = chart_fn(*args, **kwargs)
            if result is None:
                return row
            anchor = f"A{row}"
            if isinstance(result, io.BytesIO):
                from xlsx_theme import CHART_SMALL_H, CHART_HEIGHT_CM
                img = XlImage(result)
                # Scale proportionally: control height, derive width from ratio
                target_h_cm = CHART_SMALL_H if small else CHART_HEIGHT_CM
                if img.width and img.height and img.height > 0:
                    aspect = img.width / img.height
                    # 1 cm ≈ 37.8 px in Excel's coordinate system
                    target_h_px = target_h_cm * 37.8
                    target_w_px = target_h_px * aspect
                    img.width = target_w_px
                    img.height = target_h_px
                ws.add_image(img, anchor)
            else:
                # Native engine — openpyxl chart object
                ws.add_chart(result, anchor)
            rows_skip = CHART_ROWS_SMALL if small else CHART_ROWS_STANDARD
            return row + rows_skip + 1
        except Exception as e:
            logger.warning("Chart %s failed: %s", getattr(chart_fn, '__name__', '?'), e)
            return row

    def _add_heatmap(self, ws, heatmap_fn, *args, row=None, **kwargs):
        """Call a chart engine heatmap method. Returns next row.

        Native engine: writes cells directly and returns next row int.
        Matplotlib engine: returns BytesIO PNG, which we embed as an image.
        """
        if row is None:
            raise ValueError("row must be specified")
        try:
            result = heatmap_fn(*args, **kwargs)
            if result is None:
                return row
            if isinstance(result, io.BytesIO):
                from xlsx_theme import CHART_HEIGHT_CM
                img = XlImage(result)
                target_h_cm = CHART_HEIGHT_CM
                if img.width and img.height and img.height > 0:
                    aspect = img.width / img.height
                    target_h_px = target_h_cm * 37.8
                    target_w_px = target_h_px * aspect
                    img.width = target_w_px
                    img.height = target_h_px
                ws.add_image(img, f"A{row}")
                return row + CHART_ROWS_STANDARD + 1
            else:
                # Native engine — result is the next row int
                return result
        except Exception as e:
            logger.warning("Heatmap %s failed: %s", getattr(heatmap_fn, '__name__', '?'), e)
            return row

    def _write_insight(self, ws, key: str, row: int) -> int:
        """Write AI insight block with styled formatting. Returns next row."""
        text = self.insights.get(key, "")
        if not text:
            return row

        # Spacer before insight
        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        row += 1

        # Insight title row with accent background
        title_label = "🤖 AI Insight" if self.language == "en" else "🤖 AI 洞察分析"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=title_label)
        cell.font = self.styles.font_insight_title
        cell.fill = self.styles.fill_insight
        cell.alignment = self.styles.align_left
        ws.row_dimensions[row].height = ROW_HEIGHTS["h3"]
        row += 1

        # Write insight text (may be multi-line)
        lines = text.strip().split("\n")
        for line in lines:
            line = line.strip()
            if not line:
                ws.row_dimensions[row].height = 6
                row += 1
                continue
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            cell = ws.cell(row=row, column=1, value=line)
            # Bold for lines starting with emoji markers
            if line and line[0] in "📌🔍💡📈⚠":
                cell.font = Font(name=self.styles.font_insight_body.name,
                                 size=self.styles.font_insight_body.size,
                                 bold=True,
                                 color=self.styles.font_insight_title.color)
            else:
                cell.font = self.styles.font_insight_body
            cell.fill = self.styles.fill_insight
            cell.alignment = self.styles.align_wrap
            ws.row_dimensions[row].height = 30  # increased from 20pt for more room
            row += 1

        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        return row + 1

    def _write_footer(self, ws, row: int) -> int:
        """Write footer with generation timestamp."""
        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        label = (f"Generated by Customer Cross-Overview Report • {ts}"
                 if self.language == "en"
                 else f"由客户跨流程总览报告生成 • {ts}")
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = self.styles.font_footnote
        cell.alignment = self.styles.align_center
        return row + 1

    def _write_kpi_cards(self, ws, cards: list, row: int) -> int:
        """Write KPI summary cards in a row. cards = [(label, value, status), ...]
        status = 'success'|'good'|'warning'|'danger'|'neutral'
        Each card occupies 2 columns (label row + value row), up to 4 cards across.
        """
        col = 1
        value_row = row
        label_row = row + 1
        for i, (label, value, status) in enumerate(cards):
            if i > 0 and i % 4 == 0:
                # New row of cards
                value_row += 3
                label_row = value_row + 1
                col = 1
            # Merge 2 cols for value
            ws.merge_cells(start_row=value_row, start_column=col, end_row=value_row, end_column=col + 1)
            vcell = ws.cell(row=value_row, column=col, value=str(value))
            vcell.font = self.styles.font_kpi_val
            vcell.alignment = self.styles.align_center
            vcell.fill = self.styles.semantic_fill(status)
            vcell.border = self.styles.border_kpi
            # Label below
            ws.merge_cells(start_row=label_row, start_column=col, end_row=label_row, end_column=col + 1)
            lcell = ws.cell(row=label_row, column=col, value=label)
            lcell.font = self.styles.font_kpi_label
            lcell.alignment = self.styles.align_center
            lcell.border = self.styles.border_kpi
            ws.row_dimensions[value_row].height = 36
            ws.row_dimensions[label_row].height = 18
            col += 3  # gap between cards
        return label_row + 2  # spacer after cards

    # ─── Sheet builders ──────────────────────────────────────────────────

    def _build_sheet_executive(self, ws):
        """Sheet 1 — Executive Summary."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[0])

        # KPI summary table
        row = self._write_section(ws, "KPI Summary" if self.language == "en" else "KPI 概览", row)
        r = self.result
        kpis = r.kpis if isinstance(r.kpis, dict) else {}

        def _kpi_val(key, attr="current_value", default=0):
            m = kpis.get(key)
            if m is None:
                return default
            return getattr(m, attr, default)

        headers = ["KPI", "Value", "Status"]
        data = [
            ["Health Score", format_number(r.health_score), r.health_grade],
            ["Incident SLA Rate", format_pct(_kpi_val("sla_rate")), rating_text(_kpi_val("sla_rate"), self.language)],
            ["MTTR", format_duration(_kpi_val("avg_mttr") * 60, self.language), ""],
            ["Change Success Rate", format_pct(_kpi_val("change_success_rate")), rating_text(_kpi_val("change_success_rate"), self.language)],
            ["Emergency Change Rate", format_pct(_kpi_val("emergency_ratio")), ""],
            ["Request Fulfillment", format_pct(_kpi_val("fulfillment_rate")), rating_text(_kpi_val("fulfillment_rate"), self.language)],
            ["CSAT", f"{_kpi_val('request_csat'):.2f}", ""],
            ["Problem Closure Rate", format_pct(_kpi_val("problem_closure_rate")), rating_text(_kpi_val("problem_closure_rate"), self.language)],
        ]
        row = self._write_table(ws, headers, data, row, [40, COL_WIDTHS["short_text"], COL_WIDTHS["short_text"]])

        # Table of Contents with hyperlinks
        row = self._write_section(ws, "Contents" if self.language == "en" else "目录", row)
        for i, name in enumerate(names):
            cell = ws.cell(row=row, column=1, value=name)
            cell.hyperlink = f"#{name}!A1"
            cell.font = Font(color="3b82f6", underline="single", size=10)
            row += 1
        row += 1  # spacer

        # Visual Analysis section header
        row = self._write_section(ws, "Visual Analysis" if self.language == "en" else "可视化分析", row)

        # Charts: health gauge + radar (sparklines removed)
        row = self._write_chart_desc(ws, "Overall health score gauge" if self.language == "en" else "整体健康评分仪表盘", row)
        row = self._add_chart(ws, self.charts.chart_exec_health_gauge, r.health_score, row=row, small=True)
        row = self._write_chart_desc(ws, "Process health radar across four ITIL processes" if self.language == "en" else "四大ITIL流程健康雷达图", row)
        row = self._add_chart(ws, self.charts.chart_exec_process_radar, r, row=row)

        # Risk table
        if r.top_risks:
            row = self._write_section(ws, "Top Risks" if self.language == "en" else "主要风险", row)
            risk_headers = ["Priority", "Risk", "Impact", "Process"]
            risk_data = [
                [risk.priority, risk.message, risk.impact, risk.process]
                for risk in r.top_risks[:5]
            ]
            row = self._write_table(ws, risk_headers, risk_data, row,
                                    [COL_WIDTHS["short_text"], COL_WIDTHS["long_text"], COL_WIDTHS["long_text"], COL_WIDTHS["short_text"]])

        row = self._write_insight(ws, "executive_summary", row)
        self._write_footer(ws, row)

    def _build_sheet_incidents(self, ws):
        """Sheet 2 — Incident Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[1])

        # Section header for data tables
        row = self._write_section(ws, "Data Tables" if self.language == "en" else "数据表格", row)

        # Priority breakdown
        priority_rows = self.detail.priority_breakdown()
        if priority_rows:
            row = self._write_section(ws, "Priority Breakdown" if self.language == "en" else "优先级分布", row)
            headers = ["Priority", "Count", "%", "Cum%", "Avg Resolve", "Median", "Min", "Max", "Resp SLA", "Res SLA", "Violations"]
            data = [
                [p.priority, p.count, format_pct(p.pct), format_pct(p.cum_pct),
                 format_duration(p.avg_resolution_min, self.language),
                 format_duration(p.median_resolution_min, self.language),
                 format_duration(p.min_resolution_min, self.language),
                 format_duration(p.max_resolution_min, self.language),
                 format_pct(p.response_sla_rate), format_pct(p.resolution_sla_rate), p.violation_count]
                for p in priority_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Category breakdown
        category_rows = self.detail.category_breakdown()
        if category_rows:
            row = self._write_section(ws, "Category Breakdown" if self.language == "en" else "分类分布", row)
            headers = ["Category", "Count", "%", "Cum%", "Avg Resolve", "Median", "Std Dev"]
            data = [
                [c.category, c.count, format_pct(c.pct), format_pct(c.cum_pct),
                 format_duration(c.avg_resolution_min, self.language),
                 format_duration(c.median_resolution_min, self.language),
                 format_duration(c.std_resolution_min, self.language)]
                for c in category_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Visual Analysis section
        row = self._write_section(ws, "Visual Analysis" if self.language == "en" else "可视化分析", row)

        # Charts: monthly trend, priority pie, category top10 (removed mttr_boxplot, p1p2_trend)
        monthly = self.detail.monthly_trends()
        row = self._write_chart_desc(ws, "Monthly incident volume with completion rate trend" if self.language == "en" else "月度事件量与完成率趋势", row)
        row = self._add_chart(ws, self.charts.chart_inc_monthly_trend, monthly, row=row)
        row = self._write_chart_desc(ws, "Incident distribution by priority level" if self.language == "en" else "按优先级分布的事件", row)
        row = self._add_chart(ws, self.charts.chart_inc_priority_pie, priority_rows, row=row)
        row = self._write_chart_desc(ws, "Top 10 incident categories by volume" if self.language == "en" else "事件量前10的分类", row)
        row = self._add_chart(ws, self.charts.chart_inc_category_top10, category_rows, row=row)

        row = self._write_insight(ws, "incident_detail", row)
        self._write_footer(ws, row)

    def _build_sheet_sla(self, ws):
        """Sheet 3 — SLA Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[2])

        # Section header for data tables
        row = self._write_section(ws, "Data Tables" if self.language == "en" else "数据表格", row)

        # SLA summary table from result.sla_breakdown
        sla_bd = self.result.sla_breakdown or []
        if sla_bd:
            row = self._write_section(ws, "SLA Summary" if self.language == "en" else "SLA 概览", row)
            headers = ["Priority", "SLA Rate", "Total", "Compliant", "Target (h)"]
            data = []
            for s in sla_bd:
                data.append([
                    getattr(s, "priority", ""),
                    format_pct(getattr(s, "rate", 0)),
                    getattr(s, "total", 0),
                    getattr(s, "compliant", 0),
                    getattr(s, "target", ""),
                ])
            row = self._write_table(ws, headers, data, row)

        # Violation list (top 20)
        violations = self.detail.sla_violations()
        if violations:
            row = self._write_section(ws, "SLA Violations (Top 20)" if self.language == "en" else "SLA 违规 (前20)", row)
            headers = ["Order#", "Priority", "Category", "Type", "Overtime", "Resolver", "Status", "Reason"]
            data = [
                [v.order_number, v.priority, v.category, v.violation_type,
                 v.overtime, v.resolver, v.status, v.reason]
                for v in violations[:20]
            ]
            row = self._write_table(ws, headers, data, row)

        # Root cause table
        root_causes = self.detail.violation_root_causes(violations)
        if root_causes:
            row = self._write_section(ws, "Violation Root Causes" if self.language == "en" else "违规根因", row)
            headers = ["Cause", "Count", "%", "Typical Case", "Improvement"]
            data = [
                [rc.cause, rc.count, format_pct(rc.pct), rc.typical_case, rc.improvement]
                for rc in root_causes
            ]
            row = self._write_table(ws, headers, data, row)

        # Visual Analysis section
        row = self._write_section(ws, "Visual Analysis" if self.language == "en" else "可视化分析", row)

        # Charts: SLA gauge (small), monthly, violation bar, heatmap
        sla_bd = self.result.sla_breakdown or []
        overall_rate = None
        if sla_bd:
            total_compliant = sum(getattr(s, "compliant", 0) for s in sla_bd)
            total_count = sum(getattr(s, "total", 0) for s in sla_bd)
            if total_count:
                overall_rate = total_compliant / total_count
        # Use one gauge for overall SLA instead of two separate resp/res gauges
        row = self._write_chart_desc(ws, "Overall Resolution SLA compliance gauge" if self.language == "en" else "解决SLA合规仪表盘", row)
        row = self._add_chart(ws, self.charts.chart_sla_gauge_resolution, overall_rate, row=row, small=True)

        # Response SLA gauge (if data has response times)
        resp_rate = None
        if hasattr(self.detail, 'inc') and not self.detail.inc.empty:
            priority_rows_for_resp = self.detail.priority_breakdown()
            resp_rates = [getattr(p, "response_sla_rate", 0) for p in priority_rows_for_resp]
            resp_counts = [getattr(p, "count", 0) for p in priority_rows_for_resp]
            total_count = sum(resp_counts)
            if total_count > 0 and any(r > 0 for r in resp_rates):
                resp_rate = sum(r * c for r, c in zip(resp_rates, resp_counts)) / total_count
        if resp_rate is not None:
            row = self._write_chart_desc(ws, "Overall Response SLA compliance gauge" if self.language == "en" else "响应SLA合规仪表盘", row)
            row = self._add_chart(ws, self.charts.chart_sla_gauge_response, resp_rate, row=row, small=True)
        monthly = self.detail.monthly_trends()
        row = self._write_chart_desc(ws, "Monthly SLA compliance rate vs target" if self.language == "en" else "月度SLA合规率与目标对比", row)
        row = self._add_chart(ws, self.charts.chart_sla_monthly_trend, monthly, row=row)
        row = self._write_chart_desc(ws, "SLA violations breakdown by priority" if self.language == "en" else "按优先级的SLA违规分布", row)
        priority_rows = self.detail.priority_breakdown()
        row = self._add_chart(ws, self.charts.chart_sla_violation_by_priority, priority_rows, row=row)
        # Heatmap (writes cells directly)
        row = self._add_heatmap(ws, self.charts.chart_sla_violation_heatmap, ws, violations, row, row=row)

        row = self._write_insight(ws, "sla_detail", row)
        self._write_footer(ws, row)

    def _build_sheet_changes(self, ws):
        """Sheet 4 — Change Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[3])

        # Section header for data tables
        row = self._write_section(ws, "Data Tables" if self.language == "en" else "数据表格", row)

        # Change type table
        type_rows = self.detail.change_type_breakdown()
        if type_rows:
            row = self._write_section(ws, "Change Type Breakdown" if self.language == "en" else "变更类型分布", row)
            headers = ["Type", "Count", "%", "Success Rate", "Incident Rate", "Avg Duration(h)"]
            data = [
                [t.change_type, t.count, format_pct(t.pct), format_pct(t.success_rate),
                 format_pct(t.incident_rate), f"{t.avg_duration_hours:.1f}"]
                for t in type_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Change category table
        cat_rows = self.detail.change_category_breakdown()
        if cat_rows:
            row = self._write_section(ws, "Change Category Breakdown" if self.language == "en" else "变更分类分布", row)
            headers = ["Category", "Count", "Success Rate", "Failures", "Incidents", "Risk"]
            data = [
                [c.category, c.count, format_pct(c.success_rate),
                 c.failure_count, c.incident_count, c.risk_level]
                for c in cat_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Visual Analysis section
        row = self._write_section(ws, "Visual Analysis" if self.language == "en" else "可视化分析", row)

        # Charts: type pie, success trend, category bar (removed incident_scatter, planning_accuracy)
        row = self._write_chart_desc(ws, "Change distribution by type" if self.language == "en" else "按类型分布的变更", row)
        row = self._add_chart(ws, self.charts.chart_chg_type_pie, type_rows, row=row)
        monthly = self.detail.monthly_trends()
        row = self._write_chart_desc(ws, "Monthly change success rate trend" if self.language == "en" else "月度变更成功率趋势", row)
        row = self._add_chart(ws, self.charts.chart_chg_success_trend, monthly, row=row)
        row = self._write_chart_desc(ws, "Change volume and failures by category" if self.language == "en" else "按分类的变更量与失败数", row)
        row = self._add_chart(ws, self.charts.chart_chg_category_bar, cat_rows, row=row)

        row = self._write_insight(ws, "change_detail", row)
        self._write_footer(ws, row)

    def _build_sheet_requests(self, ws):
        """Sheet 5 — Request Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[4])

        # Section header for data tables
        row = self._write_section(ws, "Data Tables" if self.language == "en" else "数据表格", row)

        # Request type table
        type_rows = self.detail.request_type_breakdown()
        if type_rows:
            row = self._write_section(ws, "Request Type Breakdown" if self.language == "en" else "请求类型分布", row)
            headers = ["Type", "Count", "%", "Completion Rate", "Avg Fulfill(h)", "CSAT"]
            data = [
                [t.request_type, t.count, format_pct(t.pct), format_pct(t.completion_rate),
                 f"{t.avg_fulfillment_hours:.1f}", f"{t.csat:.2f}"]
                for t in type_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # CSAT distribution
        csat_rows = self.detail.csat_distribution()
        if csat_rows:
            row = self._write_section(ws, "CSAT Distribution" if self.language == "en" else "满意度分布", row)
            headers = ["Score", "Label", "Count", "%", "Cum%"]
            data = [
                [c.score, c.label, c.count, format_pct(c.pct), format_pct(c.cum_pct)]
                for c in csat_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Visual Analysis section
        row = self._write_section(ws, "Visual Analysis" if self.language == "en" else "可视化分析", row)

        # Charts: all 4 + heatmap kept
        row = self._write_chart_desc(ws, "Request distribution by type" if self.language == "en" else "按类型分布的请求", row)
        row = self._add_chart(ws, self.charts.chart_req_type_pie, type_rows, row=row)
        row = self._write_chart_desc(ws, "Customer satisfaction score distribution" if self.language == "en" else "客户满意度评分分布", row)
        row = self._add_chart(ws, self.charts.chart_req_csat_bar, csat_rows, row=row)
        monthly = self.detail.monthly_trends()
        row = self._write_chart_desc(ws, "Monthly request volume and average CSAT" if self.language == "en" else "月度请求量与平均满意度", row)
        row = self._add_chart(ws, self.charts.chart_req_monthly_trend, monthly, row=row)
        row = self._write_chart_desc(ws, "Fulfillment time statistics by request type" if self.language == "en" else "按请求类型的完成时间统计", row)
        row = self._add_chart(ws, self.charts.chart_req_fulfillment_bar, self.requests_df, row=row)
        # Heatmap
        row = self._add_heatmap(ws, self.charts.chart_req_dept_heatmap, ws, self.requests_df, row, row=row)

        row = self._write_insight(ws, "request_detail", row)
        self._write_footer(ws, row)

    def _build_sheet_problems(self, ws):
        """Sheet 6 — Problem Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[5])

        # Section header for data tables
        row = self._write_section(ws, "Data Tables" if self.language == "en" else "数据表格", row)

        # Problem status table
        status_rows = self.detail.problem_status_breakdown()
        if status_rows:
            row = self._write_section(ws, "Problem Status" if self.language == "en" else "问题状态", row)
            headers = ["Status", "Count", "%", "Avg Age(days)", "Suggestion"]
            data = [
                [s.status, s.count, format_pct(s.pct), f"{s.avg_age_days:.0f}", s.suggestion]
                for s in status_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Root cause category table
        rc_rows = self.detail.root_cause_category_breakdown()
        if rc_rows:
            row = self._write_section(ws, "Root Cause Categories" if self.language == "en" else "根因分类", row)
            headers = ["Category", "Count", "%", "Related Incidents", "Fix Rate", "Typical Problem"]
            data = [
                [rc.category, rc.count, format_pct(rc.pct), rc.related_incidents,
                 format_pct(rc.permanent_fix_rate), rc.typical_problem]
                for rc in rc_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Visual Analysis section
        row = self._write_section(ws, "Visual Analysis" if self.language == "en" else "可视化分析", row)

        # Charts: status funnel, rootcause pie, monthly bar (removed impact_bubble)
        row = self._write_chart_desc(ws, "Problem status funnel distribution" if self.language == "en" else "问题状态漏斗分布", row)
        row = self._add_chart(ws, self.charts.chart_prb_status_funnel, status_rows, row=row)
        row = self._write_chart_desc(ws, "Root cause category distribution" if self.language == "en" else "根因分类分布", row)
        row = self._add_chart(ws, self.charts.chart_prb_rootcause_pie, rc_rows, row=row)
        monthly = self.detail.monthly_trends()
        row = self._write_chart_desc(ws, "Monthly problem volume with cumulative trend" if self.language == "en" else "月度问题量与累计趋势", row)
        row = self._add_chart(ws, self.charts.chart_prb_monthly_bar, monthly, row=row)

        row = self._write_insight(ws, "problem_detail", row)
        self._write_footer(ws, row)

    def _build_sheet_cross_process(self, ws):
        """Sheet 7 — Cross-Process Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[6])

        # Section header for data tables
        row = self._write_section(ws, "Data Tables" if self.language == "en" else "数据表格", row)

        # Change -> Incident links
        chg_links = self.detail.change_incident_links()
        if chg_links:
            row = self._write_section(ws, "Change → Incident Links" if self.language == "en" else "变更→事件关联", row)
            headers = ["Source ID", "Type", "Target Count", "Target IDs", "Impact"]
            data = [
                [l.source_id, l.source_type, l.target_count, l.target_ids, l.impact]
                for l in chg_links[:20]
            ]
            row = self._write_table(ws, headers, data, row)

        # Visual Analysis section
        row = self._write_section(ws, "Visual Analysis" if self.language == "en" else "可视化分析", row)

        # Charts: flow bar only (removed timeline)
        row = self._write_chart_desc(ws, "Cross-process flow: changes causing incidents" if self.language == "en" else "跨流程关联: 变更导致的事件", row)
        row = self._add_chart(ws, self.charts.chart_cross_flow_bar, chg_links, None, row=row)

        row = self._write_insight(ws, "cross_process", row)
        self._write_footer(ws, row)

    def _build_sheet_personnel(self, ws):
        """Sheet 8 — Personnel & Efficiency."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[7])

        # Section header for data tables
        row = self._write_section(ws, "Data Tables" if self.language == "en" else "数据表格", row)

        # Personnel table
        personnel = self.detail.personnel_breakdown()
        if personnel:
            row = self._write_section(ws, "Personnel Performance" if self.language == "en" else "人员绩效", row)
            headers = ["Name", "Count", "%", "Avg Resolve", "Completion", "Resp SLA", "Res SLA", "Rating", "Specialty"]
            data = [
                [p.name, p.count, format_pct(p.pct),
                 format_duration(p.avg_resolution_min, self.language),
                 format_pct(p.completion_rate),
                 format_pct(p.response_sla_rate), format_pct(p.resolution_sla_rate),
                 p.rating, p.specialty]
                for p in personnel
            ]
            row = self._write_table(ws, headers, data, row)

        # Personnel Priority Distribution table (NEW)
        personnel_priority = self.detail.personnel_priority_breakdown()
        if personnel_priority:
            row = self._write_section(ws, "Personnel Priority Distribution" if self.language == "en" else "人员优先级分布", row)
            headers = ["Name", "P1", "P2", "P3", "P4", "Total", "P1+P2%"]
            data = [
                [p.name, p.p1_count, p.p2_count, p.p3_count, p.p4_count,
                 p.total, format_pct(p.high_priority_pct)]
                for p in personnel_priority[:20]  # Show top 20
            ]
            row = self._write_table(ws, headers, data, row)

        # Workload distribution
        workload = self.detail.workload_distribution(personnel)
        if workload:
            row = self._write_section(ws, "Workload Distribution" if self.language == "en" else "工作量分布", row)
            headers = ["Level", "Count", "%", "Avg Events", "Suggestion"]
            data = [
                [w.level, w.count, format_pct(w.pct), f"{w.avg_events:.1f}", w.suggestion]
                for w in workload
            ]
            row = self._write_table(ws, headers, data, row)

        # Skill coverage
        skill = self.detail.skill_coverage()
        if skill:
            row = self._write_section(ws, "Skill Coverage" if self.language == "en" else "技能覆盖", row)
            headers = ["Category", "Handlers", "Primary", "Coverage", "Risk"]
            data = [
                [s.category, s.handler_count, s.primary_handler,
                 format_pct(s.coverage_rate), s.risk_level]
                for s in skill
            ]
            row = self._write_table(ws, headers, data, row)

        # Visual Analysis section
        row = self._write_section(ws, "Visual Analysis" if self.language == "en" else "可视化分析", row)

        # Charts: top10 stacked bar (with priority), performance matrix, skill heatmap
        row = self._write_chart_desc(ws, "Top 10 personnel by ticket volume with priority breakdown" if self.language == "en" else "工单量前10的人员 (按优先级分布)", row)
        row = self._add_chart(ws, self.charts.chart_pers_top10_stacked_bar, personnel_priority, row=row)
        row = self._write_chart_desc(ws, "Performance Matrix (Volume vs MTTR)" if self.language == "en" else "绩效矩阵 (工单量 vs 平均解决时间)", row)
        row = self._add_chart(ws, self.charts.chart_pers_performance_matrix, personnel, row=row)
        # Heatmap (uses "Resolver" column)
        row = self._add_heatmap(ws, self.charts.chart_pers_skill_heatmap, ws, self.incidents_df, row, row=row)

        row = self._write_insight(ws, "personnel", row)
        self._write_footer(ws, row)

    def _build_sheet_time(self, ws):
        """Sheet 9 — Time Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[8])

        # Section header for data tables
        row = self._write_section(ws, "Data Tables" if self.language == "en" else "数据表格", row)

        # Monthly trends table
        monthly = self.detail.monthly_trends()
        if monthly:
            row = self._write_section(ws, "Monthly Trends" if self.language == "en" else "月度趋势", row)
            headers = ["Period", "Incidents", "Changes", "Requests", "Problems",
                        "Completion Rate", "Avg Resolve", "High Pri%", "MoM", "Assessment"]
            data = [
                [m.period, m.incident_count, m.change_count, m.request_count, m.problem_count,
                 format_pct(m.completion_rate),
                 format_duration(m.avg_resolution_min, self.language),
                 format_pct(m.high_priority_pct), m.mom_change, m.assessment]
                for m in monthly
            ]
            row = self._write_table(ws, headers, data, row)

        # Day-of-week table
        dow = self.detail.day_of_week_analysis()
        if dow:
            row = self._write_section(ws, "Day of Week" if self.language == "en" else "星期分布", row)
            headers = ["Day", "Count", "%", "Avg Resolve", "High Pri", "Assessment"]
            data = [
                [d.day, d.count, format_pct(d.pct),
                 format_duration(d.avg_resolution_min, self.language),
                 d.high_priority_count, d.assessment]
                for d in dow
            ]
            row = self._write_table(ws, headers, data, row)

        # Hour-of-day table
        hod = self.detail.hour_of_day_analysis()
        if hod:
            row = self._write_section(ws, "Hour of Day" if self.language == "en" else "时段分布", row)
            headers = ["Period", "Hours", "Count", "%", "Avg Resolve", "Suggestion"]
            data = [
                [h.period, h.hour_range, h.count, format_pct(h.pct),
                 format_duration(h.avg_resolution_min, self.language), h.suggestion]
                for h in hod
            ]
            row = self._write_table(ws, headers, data, row)

        # Visual Analysis section
        row = self._write_section(ws, "Visual Analysis" if self.language == "en" else "可视化分析", row)

        # Charts: four-process trend, dow bar, hour heatmap (removed quarterly, forecast)
        row = self._write_chart_desc(ws, "Four ITIL process volume trends over time" if self.language == "en" else "四大ITIL流程量时间趋势", row)
        row = self._add_chart(ws, self.charts.chart_time_four_process_trend, monthly, row=row)
        row = self._write_chart_desc(ws, "Incident volume by day of week" if self.language == "en" else "按星期分布的事件量", row)
        row = self._add_chart(ws, self.charts.chart_time_dow_bar, dow, row=row)
        # Heatmap
        row = self._add_heatmap(ws, self.charts.chart_time_hour_heatmap, ws, self.incidents_df, row, row=row)

        row = self._write_insight(ws, "time_analysis", row)
        self._write_footer(ws, row)

    def _build_sheet_actions(self, ws):
        """Sheet 10 — Action Plan."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[9])

        # Section header for data tables
        row = self._write_section(ws, "Data Tables" if self.language == "en" else "数据表格", row)

        # Build ActionPlanRow list from result.actions
        actions = self.result.actions or []
        action_plan_rows: List[ActionPlanRow] = []
        for i, a in enumerate(actions, 1):
            action_plan_rows.append(ActionPlanRow(
                seq=i,
                priority=getattr(a, "priority", "Medium"),
                action=getattr(a, "action", ""),
                source_sheet="",
                source_process=getattr(a, "process", "General"),
                responsible="",
                expected_effect=getattr(a, "expected_impact", ""),
            ))

        if action_plan_rows:
            row = self._write_section(ws, "Action Items" if self.language == "en" else "行动项", row)
            headers = ["#", "Priority", "Action", "Process", "Expected Effect"]
            data = [
                [a.seq, a.priority, a.action, a.source_process, a.expected_effect]
                for a in action_plan_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Risk summary table
        risks = self.result.risks or []
        if risks:
            row = self._write_section(ws, "Risk Summary" if self.language == "en" else "风险汇总", row)
            headers = ["ID", "Priority", "Message", "Impact", "Process"]
            data = [
                [r.id, r.priority, r.message, r.impact, r.process]
                for r in risks
            ]
            row = self._write_table(ws, headers, data, row)

        # Visual Analysis section
        row = self._write_section(ws, "Visual Analysis" if self.language == "en" else "可视化分析", row)

        # Charts: both kept
        row = self._write_chart_desc(ws, "Action items by priority level" if self.language == "en" else "按优先级的行动项", row)
        row = self._add_chart(ws, self.charts.chart_action_priority_pie, action_plan_rows, row=row)
        row = self._write_chart_desc(ws, "Action items by source process" if self.language == "en" else "按源流程的行动项", row)
        row = self._add_chart(ws, self.charts.chart_action_process_bar, action_plan_rows, row=row)

        row = self._write_insight(ws, "action_plan", row)
        self._write_footer(ws, row)

    # ─── Build & Save ────────────────────────────────────────────────────

    def build(self) -> openpyxl.Workbook:
        """Create the workbook with all 10 sheets."""
        self.wb = openpyxl.Workbook()
        if self.chart_engine_type == "matplotlib":
            self.charts = MatplotlibChartEngine(self.wb, self.language)
        else:
            self.charts = NativeChartEngine(self.wb, self.language)
        names = SHEET_NAMES[self.language]

        builders = [
            self._build_sheet_executive,
            self._build_sheet_incidents,
            self._build_sheet_sla,
            self._build_sheet_changes,
            self._build_sheet_requests,
            self._build_sheet_problems,
            self._build_sheet_cross_process,
            self._build_sheet_personnel,
            self._build_sheet_time,
            self._build_sheet_actions,
        ]

        for i, (name, builder) in enumerate(zip(names, builders)):
            if i == 0:
                ws = self.wb.active
                ws.title = name
            else:
                ws = self.wb.create_sheet(title=name)

            # Set tab color
            ws.sheet_properties.tabColor = TAB_COLORS[i]

            try:
                builder(ws)
            except Exception as e:
                logger.error("Failed to build sheet '%s': %s", name, e)
                ws.cell(row=1, column=1, value=f"Error building sheet: {e}")

        # Apply print settings
        for ws in self.wb.worksheets:
            ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
                fitToPage=True
            )
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

        return self.wb

    def save(self, filename: str = None) -> Path:
        """Build workbook, save to OUTPUT_DIR, return path."""
        if self.wb is None:
            self.build()

        if filename is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            suffix = "CN" if self.language == "zh" else "EN"
            engine_suffix = "matplotlib_chart" if self.chart_engine_type == "matplotlib" else "native_chart"
            filename = f"Comprehensive_Quality_Report_{ts}_{suffix}_{engine_suffix}.xlsx"

        output_path = OUTPUT_DIR / filename
        self.wb.save(str(output_path))

        logger.info("Report saved to %s", output_path)
        return output_path
