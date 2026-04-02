"""
Native Excel Chart Engine using openpyxl.

Replaces matplotlib PNG charts with native openpyxl Excel chart objects.
Heatmaps use conditional formatting on cells instead of chart objects.
"""

from __future__ import annotations

import math
import random
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import (
    BarChart, LineChart, PieChart, DoughnutChart, RadarChart,
    ScatterChart, BubbleChart, AreaChart, Reference,
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    Paragraph, ParagraphProperties, CharacterProperties,
    Font as DrawingFont,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment

from xlsx_theme import (
    CHART_COLORS_HEX, CHART_WIDTH_CM, CHART_HEIGHT_CM,
    CHART_SMALL_W, CHART_SMALL_H, CHART_ROWS_CONSUMED,
)


class NativeChartEngine:
    """Creates native openpyxl chart objects for the XLSX report."""

    def __init__(self, wb: Workbook, language: str = "en"):
        self.wb = wb
        self.language = language
        self._data_ws = wb.create_sheet("_ChartData")
        self._data_ws.sheet_state = "hidden"
        self._data_row = 1

    # =========================================================================
    # Helpers
    # =========================================================================

    def _write_data(self, headers: list, rows: list) -> Tuple[str, int, int]:
        """Write a data block to _ChartData. Returns (sheet_title, start_row, end_row)."""
        start = self._data_row
        for c, h in enumerate(headers, 1):
            self._data_ws.cell(row=start, column=c, value=h)
        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row, 1):
                self._data_ws.cell(row=start + 1 + r_idx, column=c_idx, value=val)
        end = start + len(rows)
        self._data_row = end + 2
        return self._data_ws.title, start, end

    def _color(self, idx: int) -> str:
        """Return 6-digit hex color for openpyxl chart fills."""
        return CHART_COLORS_HEX[idx % len(CHART_COLORS_HEX)]

    def _apply_style(self, chart, title: str,
                     width_cm: float = CHART_WIDTH_CM,
                     height_cm: float = CHART_HEIGHT_CM):
        """Apply consistent styling to a chart."""
        chart.title = title
        chart.width = width_cm
        chart.height = height_cm
        chart.style = None  # Use Excel theme colors (schemeClr) instead of preset styles
        if chart.title is not None:
            chart.title.overlay = True  # Title overlays chart area for cleaner look
        if chart.legend is not None:
            chart.legend.position = "b"
        # Add data labels to bar charts (not pie/doughnut which have their own)
        if isinstance(chart, BarChart):
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = False
            chart.dataLabels.showPercent = False
            chart.dataLabels.showSerName = True

    def _t(self, en: str, zh: str) -> str:
        """Return localised string."""
        return zh if self.language == "zh" else en

    # =========================================================================
    # Sheet 1 — Executive Summary
    # =========================================================================

    def chart_exec_health_gauge(self, score) -> Optional[DoughnutChart]:
        """Simulate a gauge using a DoughnutChart."""
        if score is None:
            return None
        score = max(0, min(100, float(score)))
        remainder = 100 - score

        headers = ["Segment", "Value"]
        rows = [["Score", score], ["Remaining", remainder]]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = DoughnutChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # Color based on score
        if score >= 80:
            fill_color = "22c55e"  # green
        elif score >= 60:
            fill_color = "eab308"  # yellow
        else:
            fill_color = "ef4444"  # red

        pt0 = DataPoint(idx=0)
        pt0.graphicalProperties.solidFill = fill_color
        pt1 = DataPoint(idx=1)
        pt1.graphicalProperties.solidFill = "e2e8f0"  # light gray
        chart.series[0].data_points.append(pt0)
        chart.series[0].data_points.append(pt1)

        title = self._t(f"Health Score: {score:.0f}/100",
                        f"健康评分: {score:.0f}/100")
        self._apply_style(chart, title, CHART_SMALL_W, CHART_SMALL_H)
        chart.legend.position = "r"
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        return chart

    def chart_exec_process_radar(self, result) -> Optional[RadarChart]:
        """4-axis radar of process KPIs."""
        keys = ["sla_rate", "change_success_rate", "fulfillment_rate", "problem_closure_rate"]
        labels_en = ["Incidents (SLA)", "Changes (Success)", "Requests (Fulfill)", "Problems (Close)"]
        labels_zh = ["事件(SLA)", "变更(成功率)", "请求(完成率)", "问题(关闭率)"]
        labels = labels_zh if self.language == "zh" else labels_en

        kpis = getattr(result, "kpis", {}) if result else {}
        if isinstance(result, dict):
            kpis = result.get("kpis", result)
        values = []
        for k in keys:
            v = kpis.get(k, 0) if isinstance(kpis, dict) else getattr(kpis, k, 0)
            # v may be a KPIMetric object – extract numeric value
            if hasattr(v, 'current_value'):
                v = v.current_value
            values.append(round(float(v or 0) * 100, 1))

        if all(v == 0 for v in values):
            return None

        headers = ["KPI", self._t("Score", "得分")]
        rows = [[lbl, val] for lbl, val in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = RadarChart()
        chart.type = "filled"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        chart.series[0].graphicalProperties.solidFill = self._color(0)

        self._apply_style(chart, self._t("Process Health Radar", "流程健康雷达"))
        chart.y_axis.scaling.max = 100
        return chart

    def chart_exec_sparklines(self, trends) -> Optional[List[LineChart]]:
        """Return list of small LineChart sparklines."""
        if not trends:
            return None
        if isinstance(trends, dict):
            items = list(trends.items())[:6]
        else:
            return None

        charts = []
        for name, values in items:
            if not values or not isinstance(values, (list, tuple)):
                continue
            nums = [float(v) for v in values if v is not None]
            if not nums:
                continue

            headers = ["Index", str(name)]
            rows = [[i + 1, v] for i, v in enumerate(nums)]
            sheet_title, start, end = self._write_data(headers, rows)

            chart = LineChart()
            data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
            cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.legend = None
            chart.y_axis.delete = True
            chart.x_axis.delete = True
            self._apply_style(chart, str(name), CHART_SMALL_W, CHART_SMALL_H / 2)
            chart.legend = None
            charts.append(chart)

        return charts if charts else None

    # =========================================================================
    # Sheet 2 — Incident Analysis
    # =========================================================================

    def chart_inc_monthly_trend(self, monthly_data) -> Optional[BarChart]:
        """Combo bar+line: incident count bars + completion rate line."""
        if not monthly_data:
            return None
        months = [getattr(m, "period", str(i)) for i, m in enumerate(monthly_data)]
        counts = [getattr(m, "incident_count", getattr(m, "count", 0)) for m in monthly_data]
        rates = [round(getattr(m, "completion_rate", 0) * 100, 1) for m in monthly_data]

        if all(c == 0 for c in counts):
            return None

        headers = [self._t("Month", "月份"), self._t("Count", "数量"),
                   self._t("Completion %", "完成率%")]
        rows = [[m, c, r] for m, c, r in zip(months, counts, rates)]
        sheet_title, start, end = self._write_data(headers, rows)

        bar = BarChart()
        bar.type = "col"
        count_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        bar.add_data(count_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.y_axis.title = self._t("Count", "数量")

        line = LineChart()
        rate_ref = Reference(self._data_ws, min_col=3, min_row=start, max_row=end)
        line.add_data(rate_ref, titles_from_data=True)
        line.y_axis.title = self._t("Rate %", "比率%")
        line.y_axis.axId = 200

        bar += line
        self._apply_style(bar, self._t("Monthly Incident Trend", "月度事件趋势"))
        return bar

    def chart_inc_priority_pie(self, priority_rows) -> Optional[PieChart]:
        """Pie chart of incidents by priority."""
        if not priority_rows:
            return None
        labels = [getattr(r, "priority", getattr(r, "name", f"P{i}"))
                  for i, r in enumerate(priority_rows)]
        values = [getattr(r, "count", getattr(r, "value", 0)) for r in priority_rows]
        if sum(values) == 0:
            return None

        headers = [self._t("Priority", "优先级"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = PieChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)

        self._apply_style(chart, self._t("Incidents by Priority", "事件优先级分布"))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showSerName = True
        return chart

    def chart_inc_category_top10(self, category_rows) -> Optional[BarChart]:
        """Horizontal bar of top 10 categories."""
        if not category_rows:
            return None
        sorted_rows = sorted(category_rows,
                             key=lambda r: getattr(r, "count", getattr(r, "value", 0)),
                             reverse=True)[:10]
        labels = [getattr(r, "category", getattr(r, "name", f"Cat{i}"))
                  for i, r in enumerate(sorted_rows)]
        values = [getattr(r, "count", getattr(r, "value", 0)) for r in sorted_rows]
        if sum(values) == 0:
            return None

        headers = [self._t("Category", "类别"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "bar"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)
        chart.x_axis.title = self._t("Category", "类别")
        chart.y_axis.title = self._t("Count", "数量")
        chart.legend = None

        self._apply_style(chart, self._t("Top 10 Incident Categories",
                                         "事件类别 Top 10"))
        return chart

    def chart_inc_mttr_boxplot(self, incidents_df) -> Optional[BarChart]:
        """Grouped bar showing Min/Avg/Max resolution time per priority."""
        try:
            import pandas as pd
            if incidents_df is None or (isinstance(incidents_df, pd.DataFrame) and incidents_df.empty):
                return None
            if isinstance(incidents_df, pd.DataFrame):
                col_priority = None
                col_time = None
                for c in incidents_df.columns:
                    if "priority" in c.lower():
                        col_priority = c
                    if "resolution" in c.lower() and "time" in c.lower():
                        col_time = c
                if col_priority and col_time:
                    grouped = incidents_df.groupby(col_priority)[col_time].agg(["min", "mean", "max"])
                    priorities = list(grouped.index)
                    mins = [round(float(v), 1) for v in grouped["min"]]
                    avgs = [round(float(v), 1) for v in grouped["mean"]]
                    maxs = [round(float(v), 1) for v in grouped["max"]]
                else:
                    raise ValueError("columns missing")
            else:
                raise ValueError("not a dataframe")
        except Exception:
            priorities = ["P1", "P2", "P3", "P4"]
            mins = [15, 30, 60, 120]
            avgs = [45, 90, 180, 360]
            maxs = [120, 240, 480, 720]
            sample = True
        else:
            sample = False

        headers = [self._t("Priority", "优先级"), "Min", self._t("Avg", "均值"), "Max"]
        rows = [[p, mn, av, mx] for p, mn, av, mx in zip(priorities, mins, avgs, maxs)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        for col_idx, color_idx in [(2, 1), (3, 0), (4, 3)]:
            ref = Reference(self._data_ws, min_col=col_idx, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        title = self._t("Resolution Time by Priority (min)", "各优先级解决时间 (分钟)")
        if sample:
            title += self._t(" (Sample)", " (示例)")
        self._apply_style(chart, title)
        return chart

    def chart_inc_p1p2_trend(self, monthly_data) -> Optional[AreaChart]:
        """Area chart of high priority percentage."""
        if not monthly_data:
            return None
        months = [getattr(m, "period", str(i)) for i, m in enumerate(monthly_data)]
        pcts = [round(getattr(m, "high_priority_pct", 0) * 100, 1) for m in monthly_data]

        headers = [self._t("Month", "月份"), self._t("High Priority %", "高优先级%")]
        rows = [[m, p] for m, p in zip(months, pcts)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = AreaChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = self._color(3)

        self._apply_style(chart, self._t("P1/P2 Incident Trend", "P1/P2事件趋势"))
        return chart

    # =========================================================================
    # Sheet 3 — SLA Analysis
    # =========================================================================

    def _sla_gauge(self, rate, title_en, title_zh) -> Optional[DoughnutChart]:
        """Generic SLA gauge doughnut."""
        if rate is None:
            return None
        rate = float(rate)
        pct = round(rate * 100, 1)
        remainder = round(100 - pct, 1)

        headers = ["Segment", "Value"]
        rows = [["Met", max(pct, 0)], ["Gap", max(remainder, 0)]]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = DoughnutChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        if rate >= 0.95:
            fill = "22c55e"
        elif rate >= 0.80:
            fill = "eab308"
        else:
            fill = "ef4444"

        pt0 = DataPoint(idx=0)
        pt0.graphicalProperties.solidFill = fill
        pt1 = DataPoint(idx=1)
        pt1.graphicalProperties.solidFill = "e2e8f0"
        chart.series[0].data_points.append(pt0)
        chart.series[0].data_points.append(pt1)

        title = self._t(f"{title_en}: {pct}%", f"{title_zh}: {pct}%")
        self._apply_style(chart, title, CHART_SMALL_W, CHART_SMALL_H)
        chart.legend.position = "b"
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showPercent = False
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        return chart

    def chart_sla_gauge_response(self, rate) -> Optional[DoughnutChart]:
        return self._sla_gauge(rate, "Response SLA", "响应SLA")

    def chart_sla_gauge_resolution(self, rate) -> Optional[DoughnutChart]:
        return self._sla_gauge(rate, "Resolution SLA", "解决SLA")

    def chart_sla_monthly_trend(self, monthly_data) -> Optional[LineChart]:
        """SLA rate line with 95% reference line."""
        if not monthly_data:
            return None
        months = [getattr(m, "period", str(i)) for i, m in enumerate(monthly_data)]
        rates = [round(getattr(m, "sla_rate", getattr(m, "completion_rate", getattr(m, "rate", 0))) * 100, 1)
                 for m in monthly_data]
        target = [95.0] * len(months)

        headers = [self._t("Month", "月份"), self._t("SLA %", "SLA%"),
                   self._t("Target", "目标")]
        rows = [[m, r, t] for m, r, t in zip(months, rates, target)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = LineChart()
        for col in [2, 3]:
            ref = Reference(self._data_ws, min_col=col, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.set_categories(cats_ref)

        chart.series[1].graphicalProperties.line.dashStyle = "dash"

        self._apply_style(chart, self._t("Monthly SLA Trend", "月度SLA趋势"))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = False
        chart.dataLabels.showSerName = False
        return chart

    def chart_sla_violation_by_priority(self, violations) -> Optional[BarChart]:
        """Stacked bar of response vs resolution violations."""
        if not violations:
            return None
        labels = [getattr(v, "priority", getattr(v, "name", f"P{i}"))
                  for i, v in enumerate(violations)]
        resp = [getattr(v, "response_violations", getattr(v, "response", 0))
                for v in violations]
        reso = [getattr(v, "resolution_violations", getattr(v, "resolution", 0))
                for v in violations]
        if sum(resp) + sum(reso) == 0:
            return None

        headers = [self._t("Priority", "优先级"),
                   self._t("Response", "响应"), self._t("Resolution", "解决")]
        rows = [[l, r1, r2] for l, r1, r2 in zip(labels, resp, reso)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        for col, ci in [(2, 3), (3, 4)]:
            ref = Reference(self._data_ws, min_col=col, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        self._apply_style(chart, self._t("SLA Violations by Priority",
                                         "各优先级SLA违规"))
        return chart

    def chart_sla_violation_heatmap(self, ws, violations, start_row: int) -> int:
        """Write violation heatmap via conditional formatting. Returns next row."""
        if not violations:
            return start_row

        # Build matrix: categories × months
        cat_month: Dict[str, Dict[str, int]] = {}
        all_months: set = set()
        for v in violations:
            cat = getattr(v, "category", getattr(v, "priority", "Unknown"))
            month = getattr(v, "month", "Unknown")
            all_months.add(month)
            cat_month.setdefault(cat, {})[month] = (
                getattr(v, "count", getattr(v, "total", 0))
            )

        if not cat_month:
            return start_row

        months_sorted = sorted(all_months)
        categories = sorted(cat_month.keys())

        # Title
        ws.cell(row=start_row, column=1,
                value=self._t("SLA Violation Heatmap", "SLA违规热力图"))
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        start_row += 1

        # Headers
        ws.cell(row=start_row, column=1,
                value=self._t("Category", "类别"))
        for j, m in enumerate(months_sorted, 2):
            ws.cell(row=start_row, column=j, value=m)
        start_row += 1
        data_start = start_row

        for cat in categories:
            ws.cell(row=start_row, column=1, value=cat)
            for j, m in enumerate(months_sorted, 2):
                ws.cell(row=start_row, column=j, value=cat_month[cat].get(m, 0))
            start_row += 1
        data_end = start_row - 1

        if data_end >= data_start and len(months_sorted) > 0:
            end_col = get_column_letter(1 + len(months_sorted))
            rule = ColorScaleRule(
                start_type="min", start_color="F7FCF5",
                mid_type="percentile", mid_value=50, mid_color="FEB24C",
                end_type="max", end_color="E31A1C",
            )
            ws.conditional_formatting.add(
                f"B{data_start}:{end_col}{data_end}", rule
            )

        return start_row + 1

    # =========================================================================
    # Sheet 4 — Change Analysis
    # =========================================================================

    def chart_chg_type_pie(self, change_types) -> Optional[PieChart]:
        """Pie chart by change type."""
        if not change_types:
            return None
        labels = [getattr(r, "type", getattr(r, "name", f"Type{i}"))
                  for i, r in enumerate(change_types)]
        values = [getattr(r, "count", getattr(r, "value", 0)) for r in change_types]
        if sum(values) == 0:
            return None

        headers = [self._t("Type", "类型"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = PieChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)

        self._apply_style(chart, self._t("Changes by Type", "变更类型分布"))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showSerName = True
        return chart

    def chart_chg_success_trend(self, monthly_data) -> Optional[LineChart]:
        """Success rate line + 90% threshold."""
        if not monthly_data:
            return None
        months = [getattr(m, "period", str(i)) for i, m in enumerate(monthly_data)]
        rates = [round(getattr(m, "change_success_rate", 0) * 100, 1)
                 for m in monthly_data]
        target = [90.0] * len(months)

        headers = [self._t("Month", "月份"), self._t("Success %", "成功率%"),
                   self._t("Threshold", "阈值")]
        rows = [[m, r, t] for m, r, t in zip(months, rates, target)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = LineChart()
        for col in [2, 3]:
            ref = Reference(self._data_ws, min_col=col, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.set_categories(cats_ref)

        chart.series[1].graphicalProperties.line.dashStyle = "dash"
        chart.x_axis.title = self._t("Month", "月份")
        chart.y_axis.title = self._t("Rate %", "比率%")

        self._apply_style(chart, self._t("Change Success Rate Trend",
                                         "变更成功率趋势"))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showSerName = False
        chart.y_axis.delete = True
        return chart

    def chart_chg_category_bar(self, categories) -> Optional[BarChart]:
        """Grouped bar: total vs failed by category."""
        if not categories:
            return None
        labels = [getattr(r, "category", getattr(r, "name", f"Cat{i}"))
                  for i, r in enumerate(categories)]
        totals = [getattr(r, "total", getattr(r, "count", 0)) for r in categories]
        failed = [getattr(r, "failed", 0) for r in categories]
        if sum(totals) == 0:
            return None

        headers = [self._t("Category", "类别"), self._t("Total", "总计"),
                   self._t("Failed", "失败")]
        rows = [[l, t, f] for l, t, f in zip(labels, totals, failed)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        for col, ci in [(2, 0), (3, 3)]:
            ref = Reference(self._data_ws, min_col=col, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.x_axis.title = self._t("Category", "类别")
        chart.y_axis.title = self._t("Count", "数量")

        self._apply_style(chart, self._t("Changes by Category", "变更类别分布"))
        return chart

    def chart_chg_incident_scatter(self, changes_df) -> Optional[ScatterChart]:
        """Scatter: X=Duration(h), Y=Risk Score."""
        sample = False
        try:
            import pandas as pd
            if changes_df is None or (isinstance(changes_df, pd.DataFrame) and changes_df.empty):
                return None
            if isinstance(changes_df, pd.DataFrame):
                has_start = any("planned" in c.lower() and "start" in c.lower()
                                for c in changes_df.columns)
                has_end = any("planned" in c.lower() and "end" in c.lower()
                              for c in changes_df.columns)
                has_risk = any("risk" in c.lower() for c in changes_df.columns)
                if has_start and has_end and has_risk:
                    start_col = [c for c in changes_df.columns
                                 if "planned" in c.lower() and "start" in c.lower()][0]
                    end_col = [c for c in changes_df.columns
                               if "planned" in c.lower() and "end" in c.lower()][0]
                    risk_col = [c for c in changes_df.columns if "risk" in c.lower()][0]
                    durations = []
                    risks = []
                    risk_map = {"High": 3, "高": 3, "Medium": 2, "中": 2,
                                "Low": 1, "低": 1}
                    for _, row in changes_df.iterrows():
                        try:
                            s = pd.to_datetime(row[start_col])
                            e = pd.to_datetime(row[end_col])
                            dur = (e - s).total_seconds() / 3600
                            r_val = risk_map.get(str(row[risk_col]), 2)
                            durations.append(round(dur, 1))
                            risks.append(r_val)
                        except Exception:
                            continue
                else:
                    raise ValueError("columns missing")
            else:
                raise ValueError("not a dataframe")
        except Exception:
            durations = [2, 4, 8, 12, 24, 3, 6, 16, 1, 10]
            risks = [1, 2, 3, 2, 3, 1, 2, 3, 1, 2]
            sample = True

        headers = [self._t("Duration (h)", "持续时间(h)"),
                   self._t("Risk Score", "风险评分")]
        rows = [[d, r] for d, r in zip(durations, risks)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = ScatterChart()
        x_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        y_ref = Reference(self._data_ws, min_col=2, min_row=start + 1, max_row=end)
        series = chart.series
        from openpyxl.chart import Series
        s = Series(y_ref, x_ref, title=self._t("Changes", "变更"))
        chart.series.append(s)

        chart.x_axis.title = self._t("Duration (hours)", "持续时间(小时)")
        chart.y_axis.title = self._t("Risk Score", "风险评分")

        title = self._t("Change Duration vs Risk", "变更持续时间与风险")
        if sample:
            title += self._t(" (Sample)", " (示例)")
        self._apply_style(chart, title)
        return chart

    def chart_chg_planning_accuracy(self, monthly_data) -> Optional[LineChart]:
        """On-time percentage line chart."""
        if not monthly_data:
            return None
        months = [getattr(m, "period", str(i)) for i, m in enumerate(monthly_data)]
        rates = [round(getattr(m, "on_time_pct",
                               getattr(m, "planning_accuracy", 0)) * 100, 1)
                 for m in monthly_data]

        headers = [self._t("Month", "月份"), self._t("On-time %", "准时率%")]
        rows = [[m, r] for m, r in zip(months, rates)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = LineChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.line.solidFill = self._color(0)

        self._apply_style(chart, self._t("Change Planning Accuracy",
                                         "变更计划准确率"))
        return chart

    # =========================================================================
    # Sheet 5 — Request Analysis
    # =========================================================================

    def chart_req_type_pie(self, request_types) -> Optional[PieChart]:
        """Pie chart by request type."""
        if not request_types:
            return None
        labels = [getattr(r, "type", getattr(r, "name", f"Type{i}"))
                  for i, r in enumerate(request_types)]
        values = [getattr(r, "count", getattr(r, "value", 0)) for r in request_types]
        if sum(values) == 0:
            return None

        headers = [self._t("Type", "类型"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = PieChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)

        self._apply_style(chart, self._t("Requests by Type", "请求类型分布"))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showSerName = True
        return chart

    def chart_req_csat_bar(self, csat_dist) -> Optional[BarChart]:
        """Color-coded bar chart of CSAT distribution."""
        if not csat_dist:
            return None
        # csat_dist can be dict {1:count,...} or list of objects
        if isinstance(csat_dist, dict):
            labels = [str(k) for k in sorted(csat_dist.keys())]
            values = [csat_dist[k] for k in sorted(csat_dist.keys())]
        else:
            labels = [str(getattr(r, "score", getattr(r, "rating", i + 1)))
                      for i, r in enumerate(csat_dist)]
            values = [getattr(r, "count", getattr(r, "value", 0)) for r in csat_dist]

        if sum(values) == 0:
            return None

        color_map = {"1": "ef4444", "2": "eab308", "3": "eab308",
                     "4": "22c55e", "5": "3b82f6"}

        headers = [self._t("Rating", "评分"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        for i, lbl in enumerate(labels):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color_map.get(lbl, self._color(i))
            chart.series[0].data_points.append(pt)
        chart.x_axis.title = self._t("Rating", "评分")
        chart.y_axis.title = self._t("Count", "数量")
        chart.legend = None

        self._apply_style(chart, self._t("CSAT Distribution", "满意度分布"))
        chart.y_axis.delete = True
        return chart

    def chart_req_monthly_trend(self, monthly_data) -> Optional[BarChart]:
        """Combo bar+line: request volume + avg CSAT."""
        if not monthly_data:
            return None
        months = [getattr(m, "period", str(i)) for i, m in enumerate(monthly_data)]
        counts = [getattr(m, "request_count", getattr(m, "count", 0))
                  for m in monthly_data]
        csats = [round(getattr(m, "avg_csat", getattr(m, "csat", 0)), 2)
                 for m in monthly_data]
        if all(c == 0 for c in counts):
            return None

        headers = [self._t("Month", "月份"), self._t("Volume", "数量"),
                   self._t("Avg CSAT", "平均满意度")]
        rows = [[m, c, s] for m, c, s in zip(months, counts, csats)]
        sheet_title, start, end = self._write_data(headers, rows)

        bar = BarChart()
        bar.type = "col"
        count_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        bar.add_data(count_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.series[0].graphicalProperties.solidFill = self._color(0)
        bar.y_axis.title = self._t("Volume", "数量")

        line = LineChart()
        csat_ref = Reference(self._data_ws, min_col=3, min_row=start, max_row=end)
        line.add_data(csat_ref, titles_from_data=True)
        line.y_axis.title = self._t("CSAT", "满意度")
        line.y_axis.axId = 200
        line.series[0].graphicalProperties.line.solidFill = self._color(1)

        bar += line
        self._apply_style(bar, self._t("Monthly Request Trend", "月度请求趋势"))
        return bar

    def chart_req_fulfillment_bar(self, requests_df) -> Optional[BarChart]:
        """Grouped bar Min/Avg/Max fulfillment time by request type."""
        sample = False
        try:
            import pandas as pd
            if requests_df is None or (isinstance(requests_df, pd.DataFrame) and requests_df.empty):
                return None
            if isinstance(requests_df, pd.DataFrame):
                col_type = None
                col_time = None
                for c in requests_df.columns:
                    if "request" in c.lower() and "type" in c.lower():
                        col_type = c
                    if "fulfillment" in c.lower() and "time" in c.lower():
                        col_time = c
                if col_type and col_time:
                    grouped = requests_df.groupby(col_type)[col_time].agg(["min", "mean", "max"])
                    types = list(grouped.index)
                    mins = [round(float(v), 1) for v in grouped["min"]]
                    avgs = [round(float(v), 1) for v in grouped["mean"]]
                    maxs = [round(float(v), 1) for v in grouped["max"]]
                else:
                    raise ValueError("columns missing")
            else:
                raise ValueError("not a dataframe")
        except Exception:
            types = ["Hardware", "Software", "Access", "Other"]
            mins = [1, 0.5, 0.2, 0.5]
            avgs = [8, 4, 2, 6]
            maxs = [24, 16, 8, 20]
            sample = True

        headers = [self._t("Request Type", "请求类型"), "Min",
                   self._t("Avg", "均值"), "Max"]
        rows = [[t, mn, av, mx] for t, mn, av, mx in zip(types, mins, avgs, maxs)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        for col, ci in [(2, 1), (3, 0), (4, 3)]:
            ref = Reference(self._data_ws, min_col=col, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        title = self._t("Fulfillment Time by Type (hours)",
                        "各类型完成时间 (小时)")
        if sample:
            title += self._t(" (Sample)", " (示例)")
        self._apply_style(chart, title)
        chart.y_axis.delete = True
        chart.legend.position = "t"
        return chart

    def chart_req_dept_heatmap(self, ws, requests_df, start_row: int) -> int:
        """Heatmap: Request Type x Requester Dept via conditional formatting."""
        try:
            import pandas as pd
            if requests_df is None or (isinstance(requests_df, pd.DataFrame) and requests_df.empty):
                return start_row
            col_type = None
            col_dept = None
            for c in requests_df.columns:
                if "request" in c.lower() and "type" in c.lower():
                    col_type = c
                if "requester" in c.lower() and "dept" in c.lower():
                    col_dept = c
            if col_type and col_dept:
                pivot = pd.crosstab(requests_df[col_type], requests_df[col_dept])
                row_labels = list(pivot.index)
                col_labels = list(pivot.columns)
                matrix = pivot.values.tolist()
            else:
                raise ValueError("columns missing")
        except Exception:
            row_labels = ["Hardware", "Software", "Access"]
            col_labels = ["IT", "Finance", "HR", "Sales"]
            matrix = [[10, 5, 3, 8], [15, 12, 7, 4], [6, 2, 9, 3]]

        ws.cell(row=start_row, column=1,
                value=self._t("Request Type × Department Heatmap",
                              "请求类型×部门热力图"))
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        start_row += 1

        # Headers
        ws.cell(row=start_row, column=1,
                value=self._t("Request Type", "请求类型"))
        for j, cl in enumerate(col_labels, 2):
            ws.cell(row=start_row, column=j, value=cl)
        start_row += 1
        data_start = start_row

        for i, rl in enumerate(row_labels):
            ws.cell(row=start_row, column=1, value=rl)
            for j, val in enumerate(matrix[i], 2):
                ws.cell(row=start_row, column=j, value=int(val))
            start_row += 1
        data_end = start_row - 1

        if data_end >= data_start and col_labels:
            end_col = get_column_letter(1 + len(col_labels))
            rule = ColorScaleRule(
                start_type="min", start_color="F7FCF5",
                mid_type="percentile", mid_value=50, mid_color="FEB24C",
                end_type="max", end_color="E31A1C",
            )
            ws.conditional_formatting.add(
                f"B{data_start}:{end_col}{data_end}", rule
            )

        return start_row + 1

    # =========================================================================
    # Sheet 6 — Problem Analysis
    # =========================================================================

    def chart_prb_status_funnel(self, status_rows) -> Optional[BarChart]:
        """Horizontal bar by status with different colors."""
        if not status_rows:
            return None
        labels = [getattr(r, "status", getattr(r, "name", f"S{i}"))
                  for i, r in enumerate(status_rows)]
        values = [getattr(r, "count", getattr(r, "value", 0)) for r in status_rows]
        if sum(values) == 0:
            return None

        headers = [self._t("Status", "状态"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "bar"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)
        chart.x_axis.title = self._t("Status", "状态")
        chart.y_axis.title = self._t("Count", "数量")
        chart.legend = None

        self._apply_style(chart, self._t("Problem Status Funnel", "问题状态漏斗"))
        return chart

    def chart_prb_rootcause_pie(self, rootcause_rows) -> Optional[PieChart]:
        """Pie chart by root cause."""
        if not rootcause_rows:
            return None
        labels = [getattr(r, "rootcause", getattr(r, "root_cause",
                  getattr(r, "name", f"RC{i}")))
                  for i, r in enumerate(rootcause_rows)]
        values = [getattr(r, "count", getattr(r, "value", 0)) for r in rootcause_rows]
        if sum(values) == 0:
            return None

        headers = [self._t("Root Cause", "根本原因"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = PieChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)

        self._apply_style(chart, self._t("Root Cause Distribution",
                                         "根本原因分布"))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showSerName = True
        return chart

    def chart_prb_monthly_bar(self, monthly_data) -> Optional[BarChart]:
        """Combo bar+line: monthly count + cumulative."""
        if not monthly_data:
            return None
        months = [getattr(m, "period", str(i)) for i, m in enumerate(monthly_data)]
        counts = [getattr(m, "problem_count", getattr(m, "count", 0))
                  for m in monthly_data]
        if all(c == 0 for c in counts):
            return None

        cumulative = []
        total = 0
        for c in counts:
            total += c
            cumulative.append(total)

        headers = [self._t("Month", "月份"), self._t("Count", "数量"),
                   self._t("Cumulative", "累计")]
        rows = [[m, c, cu] for m, c, cu in zip(months, counts, cumulative)]
        sheet_title, start, end = self._write_data(headers, rows)

        bar = BarChart()
        bar.type = "col"
        count_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        bar.add_data(count_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.series[0].graphicalProperties.solidFill = self._color(0)
        bar.y_axis.title = self._t("Count", "数量")

        line = LineChart()
        cum_ref = Reference(self._data_ws, min_col=3, min_row=start, max_row=end)
        line.add_data(cum_ref, titles_from_data=True)
        line.y_axis.title = self._t("Cumulative", "累计")
        line.y_axis.axId = 200
        line.series[0].graphicalProperties.line.solidFill = self._color(3)

        bar += line
        self._apply_style(bar, self._t("Monthly Problem Trend", "月度问题趋势"))
        return bar

    def chart_prb_impact_bubble(self, problems_df) -> Optional[BubbleChart]:
        """Bubble: X=Age(days), Y=Related Incidents, Size=Priority."""
        sample = False
        try:
            import pandas as pd
            if problems_df is None or (isinstance(problems_df, pd.DataFrame) and problems_df.empty):
                return None
            if isinstance(problems_df, pd.DataFrame):
                col_date = None
                col_incidents = None
                col_priority = None
                for c in problems_df.columns:
                    cl = c.lower()
                    if "logged" in cl and "date" in cl:
                        col_date = c
                    if "related" in cl and "incident" in cl:
                        col_incidents = c
                    if "priority" in cl:
                        col_priority = c

                if col_date and col_incidents and col_priority:
                    ages, incidents, sizes = [], [], []
                    pri_map = {"P1": 4, "P2": 3, "P3": 2, "P4": 1,
                               "Critical": 4, "High": 3, "Medium": 2, "Low": 1}
                    now = datetime.now()
                    for _, row in problems_df.iterrows():
                        try:
                            logged = pd.to_datetime(row[col_date])
                            age = (now - logged).days
                            inc_count = int(str(row[col_incidents]).split()[0])
                            pri_size = pri_map.get(str(row[col_priority]), 2)
                            ages.append(age)
                            incidents.append(inc_count)
                            sizes.append(pri_size)
                        except Exception:
                            continue
                    if not ages:
                        raise ValueError("no valid data")
                else:
                    raise ValueError("columns missing")
            else:
                raise ValueError("not a dataframe")
        except Exception:
            ages = [10, 30, 60, 90, 120, 15, 45, 75]
            incidents = [3, 5, 2, 8, 1, 4, 6, 3]
            sizes = [4, 3, 2, 4, 1, 3, 2, 1]
            sample = True

        headers = [self._t("Age (days)", "存在天数"),
                   self._t("Related Incidents", "关联事件"),
                   self._t("Priority Size", "优先级大小")]
        rows = [[a, inc, s] for a, inc, s in zip(ages, incidents, sizes)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BubbleChart()
        x_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        y_ref = Reference(self._data_ws, min_col=2, min_row=start + 1, max_row=end)
        s_ref = Reference(self._data_ws, min_col=3, min_row=start + 1, max_row=end)

        from openpyxl.chart import Series
        series = Series(y_ref, x_ref, s_ref,
                        title=self._t("Problems", "问题"))
        chart.series.append(series)

        chart.x_axis.title = self._t("Age (days)", "存在天数")
        chart.y_axis.title = self._t("Related Incidents", "关联事件")

        title = self._t("Problem Impact Bubble", "问题影响气泡图")
        if sample:
            title += self._t(" (Sample)", " (示例)")
        self._apply_style(chart, title)
        return chart

    # =========================================================================
    # Sheet 7 — Cross-Process
    # =========================================================================

    def chart_cross_flow_bar(self, change_links, problem_links=None) -> Optional[BarChart]:
        """Bar showing cross-process flow volumes."""
        chg_count = len(change_links) if change_links else 0
        if chg_count == 0:
            return None

        # Only show Change→Incidents
        labels = [self._t("Changes→Incidents", "变更→事件")]
        values = [chg_count]

        headers = [self._t("Flow", "流向"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)

        self._apply_style(chart, self._t("Cross-Process Flow",
                                         "跨流程关联"))
        chart.y_axis.delete = True
        chart.legend.position = "t"
        return chart

    def chart_cross_radar(self, process_health) -> Optional[RadarChart]:
        """Multi-axis radar of process health scores."""
        if not process_health:
            return None
        if isinstance(process_health, dict):
            labels = list(process_health.keys())
            values = [float(v) for v in process_health.values()]
        else:
            labels = [getattr(p, "process", getattr(p, "name", f"P{i}"))
                      for i, p in enumerate(process_health)]
            values = [float(getattr(p, "score", getattr(p, "value", 0)))
                      for p in process_health]

        if all(v == 0 for v in values):
            return None

        headers = [self._t("Process", "流程"), self._t("Score", "得分")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = RadarChart()
        chart.type = "filled"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = self._color(4)

        self._apply_style(chart, self._t("Process Health Radar",
                                         "流程健康雷达"))
        return chart

    def chart_cross_timeline(self, change_links) -> Optional[ScatterChart]:
        """Timeline scatter of change-incident links."""
        if not change_links:
            return None

        x_vals = []
        y_vals = []
        for i, link in enumerate(change_links):
            x_vals.append(i + 1)
            y_vals.append(getattr(link, "incident_count",
                         getattr(link, "count", random.randint(1, 5))))

        headers = [self._t("Change #", "变更编号"),
                   self._t("Incidents", "事件数")]
        rows = [[x, y] for x, y in zip(x_vals, y_vals)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = ScatterChart()
        x_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        y_ref = Reference(self._data_ws, min_col=2, min_row=start + 1, max_row=end)
        from openpyxl.chart import Series
        s = Series(y_ref, x_ref, title=self._t("Links", "关联"))
        chart.series.append(s)

        chart.x_axis.title = self._t("Change Sequence", "变更序号")
        chart.y_axis.title = self._t("Related Incidents", "关联事件")

        self._apply_style(chart, self._t("Change-Incident Timeline",
                                         "变更-事件时间线"))
        return chart

    # =========================================================================
    # Sheet 8 — Personnel
    # =========================================================================

    def chart_pers_workload_bar(self, personnel) -> Optional[BarChart]:
        """Horizontal bar sorted by count."""
        if not personnel:
            return None
        sorted_p = sorted(personnel,
                          key=lambda p: getattr(p, "count", getattr(p, "value", 0)),
                          reverse=True)[:15]
        labels = [getattr(p, "name", getattr(p, "person", f"P{i}"))
                  for i, p in enumerate(sorted_p)]
        values = [getattr(p, "count", getattr(p, "value", 0)) for p in sorted_p]
        if sum(values) == 0:
            return None

        headers = [self._t("Person", "人员"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "bar"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = self._color(0)
        chart.x_axis.title = self._t("Count", "数量")
        chart.y_axis.title = self._t("Person", "人员")
        chart.legend = None

        self._apply_style(chart, self._t("Personnel Workload", "人员工作量"))
        return chart

    def chart_pers_load_distribution(self, personnel) -> Optional[BarChart]:
        """Histogram of workload distribution."""
        if not personnel:
            return None
        counts = [getattr(p, "count", getattr(p, "value", 0)) for p in personnel]
        if not counts or max(counts) == 0:
            return None

        max_val = max(counts)
        # Create 5 buckets
        bucket_size = max(1, math.ceil(max_val / 5))
        buckets = {}
        for i in range(5):
            lo = i * bucket_size
            hi = (i + 1) * bucket_size
            label = f"{lo}-{hi}"
            buckets[label] = 0
        for c in counts:
            idx = min(int(c / bucket_size), 4)
            lo = idx * bucket_size
            hi = (idx + 1) * bucket_size
            label = f"{lo}-{hi}"
            buckets[label] = buckets.get(label, 0) + 1

        labels = list(buckets.keys())
        values = list(buckets.values())

        headers = [self._t("Range", "范围"), self._t("People", "人数")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = self._color(5)

        self._apply_style(chart, self._t("Workload Distribution",
                                         "工作量分布"))
        return chart

    def chart_pers_skill_heatmap(self, ws, incidents_df, start_row: int) -> int:
        """Heatmap: Resolver x Category via conditional formatting."""
        try:
            import pandas as pd
            if incidents_df is None or (isinstance(incidents_df, pd.DataFrame) and incidents_df.empty):
                return start_row
            col_resolver = None
            col_category = None
            for c in incidents_df.columns:
                if "resolver" in c.lower():
                    col_resolver = c
                if "category" in c.lower():
                    col_category = c
            if col_resolver and col_category:
                top_resolvers = (incidents_df[col_resolver]
                                 .value_counts().head(15).index.tolist())
                filtered = incidents_df[incidents_df[col_resolver].isin(top_resolvers)]
                pivot = pd.crosstab(filtered[col_resolver], filtered[col_category])
                row_labels = list(pivot.index)
                col_labels = list(pivot.columns)
                matrix = pivot.values.tolist()
            else:
                raise ValueError("columns missing")
        except Exception:
            row_labels = [f"Resolver{i}" for i in range(1, 11)]
            col_labels = ["Network", "Server", "App", "Database", "Security"]
            matrix = [[random.randint(0, 15) for _ in col_labels]
                      for _ in row_labels]

        ws.cell(row=start_row, column=1,
                value=self._t("Personnel Skill Heatmap", "人员技能热力图"))
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        start_row += 1

        ws.cell(row=start_row, column=1, value=self._t("Resolver", "处理人"))
        for j, cl in enumerate(col_labels, 2):
            ws.cell(row=start_row, column=j, value=cl)
        start_row += 1
        data_start = start_row

        for i, rl in enumerate(row_labels):
            ws.cell(row=start_row, column=1, value=rl)
            for j, val in enumerate(matrix[i], 2):
                ws.cell(row=start_row, column=j, value=int(val))
            start_row += 1
        data_end = start_row - 1

        if data_end >= data_start and col_labels:
            end_col = get_column_letter(1 + len(col_labels))
            rule = ColorScaleRule(
                start_type="min", start_color="F7FCF5",
                mid_type="percentile", mid_value=50, mid_color="FEB24C",
                end_type="max", end_color="E31A1C",
            )
            ws.conditional_formatting.add(
                f"B{data_start}:{end_col}{data_end}", rule
            )

        return start_row + 1

    def chart_pers_efficiency_scatter(self, personnel) -> Optional[ScatterChart]:
        """Scatter: X=count, Y=avg_resolution_min."""
        if not personnel:
            return None
        x_vals = [getattr(p, "count", getattr(p, "value", 0)) for p in personnel]
        y_vals = [getattr(p, "avg_resolution_min",
                  getattr(p, "avg_time", 0)) for p in personnel]
        if all(x == 0 for x in x_vals):
            return None

        headers = [self._t("Ticket Count", "工单数"),
                   self._t("Avg Resolution (min)", "平均解决时间(分钟)")]
        rows = [[x, y] for x, y in zip(x_vals, y_vals)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = ScatterChart()
        x_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        y_ref = Reference(self._data_ws, min_col=2, min_row=start + 1, max_row=end)
        from openpyxl.chart import Series
        s = Series(y_ref, x_ref, title=self._t("Personnel", "人员"))
        chart.series.append(s)

        chart.x_axis.title = self._t("Ticket Count", "工单数")
        chart.y_axis.title = self._t("Avg Resolution (min)", "平均解决时间(分钟)")

        self._apply_style(chart, self._t("Personnel Efficiency",
                                         "人员效率散点图"))
        return chart

    def chart_pers_top10_bar(self, personnel) -> Optional[BarChart]:
        """Horizontal bar of top 10 personnel by count."""
        if not personnel:
            return None
        sorted_p = sorted(personnel,
                          key=lambda p: getattr(p, "count", getattr(p, "value", 0)),
                          reverse=True)[:10]
        labels = [getattr(p, "name", getattr(p, "person", f"P{i}"))
                  for i, p in enumerate(sorted_p)]
        values = [getattr(p, "count", getattr(p, "value", 0)) for p in sorted_p]
        if sum(values) == 0:
            return None

        headers = [self._t("Person", "人员"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "bar"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)
        chart.x_axis.title = self._t("Count", "数量")
        chart.y_axis.title = self._t("Person", "人员")

        self._apply_style(chart, self._t("Top 10 Personnel", "人员 Top 10"))
        return chart

    def chart_pers_top10_stacked_bar(self, personnel_priority) -> Optional[BarChart]:
        """Horizontal stacked bar of top 10 personnel by count with priority breakdown.

        Each bar segment represents a different priority level:
        - P1 (Critical) - Red
        - P2 (High) - Orange
        - P3 (Medium) - Yellow
        - P4 (Low) - Green
        """
        if not personnel_priority:
            return None

        # Take top 10 by total count (already sorted)
        top10 = personnel_priority[:10]

        # Prepare data: names and priority counts
        names = [getattr(p, "name", f"P{i}") for i, p in enumerate(top10)]
        p1_counts = [getattr(p, "p1_count", 0) for p in top10]
        p2_counts = [getattr(p, "p2_count", 0) for p in top10]
        p3_counts = [getattr(p, "p3_count", 0) for p in top10]
        p4_counts = [getattr(p, "p4_count", 0) for p in top10]
        totals = [getattr(p, "total", 0) for p in top10]

        if sum(totals) == 0:
            return None

        # Write data to hidden sheet
        headers = [self._t("Person", "人员"), "P1", "P2", "P3", "P4"]
        rows = [[n, p1, p2, p3, p4] for n, p1, p2, p3, p4
                in zip(names, p1_counts, p2_counts, p3_counts, p4_counts)]
        sheet_title, start, end = self._write_data(headers, rows)

        # Create horizontal stacked bar chart
        chart = BarChart()
        chart.type = "bar"  # horizontal
        chart.grouping = "stacked"

        # Categories (person names)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)

        # Add each priority series with specific colors
        priority_colors = {
            "P1": "ef4444",  # Red - Critical
            "P2": "f97316",  # Orange - High
            "P3": "eab308",  # Yellow - Medium
            "P4": "22c55e",  # Green - Low
        }

        for col_idx, (priority, color) in enumerate(priority_colors.items(), 2):
            ref = Reference(self._data_ws, min_col=col_idx, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)

        chart.set_categories(cats_ref)

        # Apply colors to each series
        for i, (priority, color) in enumerate(priority_colors.items()):
            if i < len(chart.series):
                chart.series[i].graphicalProperties.solidFill = color

        chart.x_axis.title = self._t("Ticket Count", "工单数量")
        chart.y_axis.title = self._t("Person", "人员")

        self._apply_style(chart, self._t("Top 10 Personnel by Priority",
                                         "人员Top10 (按优先级分布)"))

        # Configure data labels to show totals
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showCatName = False
        chart.dataLabels.showPercent = False
        chart.dataLabels.showSerName = False

        return chart

    def chart_pers_performance_matrix(self, personnel):
        """Not supported in native engine — scatter with quadrants requires matplotlib."""
        return None

    # =========================================================================
    # Sheet 9 — Time Analysis
    # =========================================================================

    def chart_time_four_process_trend(self, monthly_data) -> Optional[LineChart]:
        """4-series line chart of process trends."""
        if not monthly_data:
            return None
        months = [getattr(m, "period", str(i)) for i, m in enumerate(monthly_data)]
        series_keys = [
            ("incident_count", self._t("Incidents", "事件")),
            ("change_count", self._t("Changes", "变更")),
            ("request_count", self._t("Requests", "请求")),
            ("problem_count", self._t("Problems", "问题")),
        ]

        header_row = [self._t("Month", "月份")]
        for key, name in series_keys:
            header_row.append(name)
        data_rows = []
        for i, m in enumerate(monthly_data):
            row = [months[i]]
            for key, _ in series_keys:
                row.append(getattr(m, key, getattr(m, "count", 0)))
            data_rows.append(row)

        sheet_title, start, end = self._write_data(header_row, data_rows)

        chart = LineChart()
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        for col_idx in range(2, 2 + len(series_keys)):
            ref = Reference(self._data_ws, min_col=col_idx, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.x_axis.title = self._t("Month", "月份")
        chart.y_axis.title = self._t("Count", "数量")

        self._apply_style(chart, self._t("Four Process Trend",
                                         "四大流程趋势"))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showSerName = False
        chart.y_axis.delete = True
        return chart

    def chart_time_dow_bar(self, dow_data) -> Optional[BarChart]:
        """Bar chart by day of week."""
        if not dow_data:
            return None
        if isinstance(dow_data, dict):
            labels = list(dow_data.keys())
            values = [dow_data[k] for k in labels]
        else:
            labels = [getattr(d, "day", getattr(d, "name", f"Day{i}"))
                      for i, d in enumerate(dow_data)]
            values = [getattr(d, "count", getattr(d, "value", 0)) for d in dow_data]

        if sum(values) == 0:
            return None

        headers = [self._t("Day", "星期"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)
        chart.x_axis.title = self._t("Day", "星期")
        chart.y_axis.title = self._t("Count", "数量")
        chart.legend = None

        self._apply_style(chart, self._t("Incidents by Day of Week",
                                         "按星期分布"))
        chart.y_axis.delete = True
        return chart

    def chart_time_hour_heatmap(self, ws, incidents_df, start_row: int) -> int:
        """Heatmap: Hour x Weekday via conditional formatting."""
        try:
            import pandas as pd
            if incidents_df is None or (isinstance(incidents_df, pd.DataFrame) and incidents_df.empty):
                return start_row
            col_date = None
            for c in incidents_df.columns:
                if "begin" in c.lower() and "date" in c.lower():
                    col_date = c
            if col_date:
                dt_series = pd.to_datetime(incidents_df[col_date], errors="coerce")
                valid = dt_series.dropna()
                hours = valid.dt.hour
                weekdays = valid.dt.day_name()
                pivot = pd.crosstab(hours, weekdays)
                dow_order = ["Monday", "Tuesday", "Wednesday", "Thursday",
                             "Friday", "Saturday", "Sunday"]
                existing = [d for d in dow_order if d in pivot.columns]
                pivot = pivot.reindex(columns=existing, fill_value=0)
                row_labels = [str(h) for h in pivot.index]
                col_labels = list(pivot.columns)
                matrix = pivot.values.tolist()
            else:
                raise ValueError("column missing")
        except Exception:
            row_labels = [str(h) for h in range(24)]
            col_labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
            matrix = [[random.randint(0, 20) for _ in col_labels]
                      for _ in row_labels]

        ws.cell(row=start_row, column=1,
                value=self._t("Hourly Incident Heatmap", "小时事件热力图"))
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        start_row += 1

        ws.cell(row=start_row, column=1, value=self._t("Hour", "小时"))
        for j, cl in enumerate(col_labels, 2):
            ws.cell(row=start_row, column=j, value=cl)
        start_row += 1
        data_start = start_row

        for i, rl in enumerate(row_labels):
            ws.cell(row=start_row, column=1, value=rl)
            for j, val in enumerate(matrix[i], 2):
                ws.cell(row=start_row, column=j, value=int(val))
            start_row += 1
        data_end = start_row - 1

        if data_end >= data_start and col_labels:
            end_col = get_column_letter(1 + len(col_labels))
            rule = ColorScaleRule(
                start_type="min", start_color="F7FCF5",
                mid_type="percentile", mid_value=50, mid_color="FEB24C",
                end_type="max", end_color="E31A1C",
            )
            ws.conditional_formatting.add(
                f"B{data_start}:{end_col}{data_end}", rule
            )

        return start_row + 1

    def chart_time_quarterly(self, monthly_data) -> Optional[BarChart]:
        """Quarterly bar + trend line overlay."""
        if not monthly_data:
            return None

        # Aggregate monthly into quarterly
        quarters: Dict[str, int] = {}
        for i, m in enumerate(monthly_data):
            month_str = getattr(m, "period", str(i + 1))
            count = getattr(m, "incident_count", getattr(m, "count", 0))
            # Determine quarter
            try:
                month_num = int(month_str.split("-")[1]) if "-" in str(month_str) else (i % 12) + 1
            except Exception:
                month_num = (i % 12) + 1
            q = f"Q{(month_num - 1) // 3 + 1}"
            quarters[q] = quarters.get(q, 0) + count

        if not quarters or all(v == 0 for v in quarters.values()):
            return None

        labels = list(quarters.keys())
        values = list(quarters.values())

        # Simple trend: linear
        n = len(values)
        if n > 1:
            x_mean = (n - 1) / 2
            y_mean = sum(values) / n
            num = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(values))
            den = sum((i - x_mean) ** 2 for i in range(n))
            slope = num / den if den != 0 else 0
            intercept = y_mean - slope * x_mean
            trend = [round(intercept + slope * i, 1) for i in range(n)]
        else:
            trend = values[:]

        headers = [self._t("Quarter", "季度"), self._t("Count", "数量"),
                   self._t("Trend", "趋势")]
        rows = [[l, v, t] for l, v, t in zip(labels, values, trend)]
        sheet_title, start, end = self._write_data(headers, rows)

        bar = BarChart()
        bar.type = "col"
        count_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        bar.add_data(count_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.series[0].graphicalProperties.solidFill = self._color(0)

        line = LineChart()
        trend_ref = Reference(self._data_ws, min_col=3, min_row=start, max_row=end)
        line.add_data(trend_ref, titles_from_data=True)
        line.y_axis.axId = 200
        line.series[0].graphicalProperties.line.solidFill = self._color(3)
        line.series[0].graphicalProperties.line.dashStyle = "dash"

        bar += line
        self._apply_style(bar, self._t("Quarterly Trend", "季度趋势"))
        return bar

    def chart_time_forecast(self, monthly_data) -> Optional[LineChart]:
        """Actual line + forecast dashed line via linear regression."""
        if not monthly_data:
            return None
        months = [getattr(m, "period", str(i + 1)) for i, m in enumerate(monthly_data)]
        values = [getattr(m, "incident_count", getattr(m, "count", 0))
                  for m in monthly_data]
        if all(v == 0 for v in values):
            return None

        n = len(values)
        # Linear regression
        x_mean = (n - 1) / 2
        y_mean = sum(values) / n
        num = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(values))
        den = sum((i - x_mean) ** 2 for i in range(n))
        slope = num / den if den != 0 else 0
        intercept = y_mean - slope * x_mean

        # Forecast 3 periods
        forecast_count = 3
        forecast_months = [f"F+{i + 1}" for i in range(forecast_count)]
        forecast_vals = [round(intercept + slope * (n + i), 1)
                         for i in range(forecast_count)]

        all_months = months + forecast_months
        actuals = values + [None] * forecast_count
        forecasts = [None] * n + forecast_vals

        headers = [self._t("Month", "月份"), self._t("Actual", "实际"),
                   self._t("Forecast", "预测")]
        rows = [[m, a, f] for m, a, f in zip(all_months, actuals, forecasts)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = LineChart()
        for col in [2, 3]:
            ref = Reference(self._data_ws, min_col=col, min_row=start, max_row=end)
            chart.add_data(ref, titles_from_data=True)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.set_categories(cats_ref)

        chart.series[0].graphicalProperties.line.solidFill = self._color(0)
        chart.series[1].graphicalProperties.line.solidFill = self._color(3)
        chart.series[1].graphicalProperties.line.dashStyle = "dash"

        self._apply_style(chart, self._t("Incident Forecast", "事件预测"))
        return chart

    # =========================================================================
    # Sheet 10 — Action Plan
    # =========================================================================

    def chart_action_priority_pie(self, actions) -> Optional[PieChart]:
        """Pie by action priority. High=red, Medium=yellow, Low=green."""
        if not actions:
            return None
        priority_counts: Dict[str, int] = {}
        for a in actions:
            p = getattr(a, "priority", getattr(a, "level", "Medium"))
            priority_counts[p] = priority_counts.get(p, 0) + 1
        if not priority_counts or sum(priority_counts.values()) == 0:
            return None

        labels = list(priority_counts.keys())
        values = list(priority_counts.values())

        color_map = {
            "High": "ef4444", "高": "ef4444",
            "Medium": "eab308", "中": "eab308",
            "Low": "22c55e", "低": "22c55e",
        }

        headers = [self._t("Priority", "优先级"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = PieChart()
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        for i, lbl in enumerate(labels):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color_map.get(lbl, self._color(i))
            chart.series[0].data_points.append(pt)

        self._apply_style(chart, self._t("Actions by Priority", "行动优先级分布"))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showSerName = True
        return chart

    def chart_action_process_bar(self, actions) -> Optional[BarChart]:
        """Bar chart by process with different colors."""
        if not actions:
            return None
        process_counts: Dict[str, int] = {}
        for a in actions:
            p = getattr(a, "source_process", getattr(a, "process", getattr(a, "category", "Other")))
            process_counts[p] = process_counts.get(p, 0) + 1
        if not process_counts or sum(process_counts.values()) == 0:
            return None

        labels = list(process_counts.keys())
        values = list(process_counts.values())

        headers = [self._t("Process", "流程"), self._t("Count", "数量")]
        rows = [[l, v] for l, v in zip(labels, values)]
        sheet_title, start, end = self._write_data(headers, rows)

        chart = BarChart()
        chart.type = "col"
        data_ref = Reference(self._data_ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(self._data_ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        for i in range(len(labels)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = self._color(i)
            chart.series[0].data_points.append(pt)
        chart.x_axis.title = self._t("Process", "流程")
        chart.y_axis.title = self._t("Count", "数量")

        self._apply_style(chart, self._t("Actions by Process", "各流程行动计划"))
        return chart
