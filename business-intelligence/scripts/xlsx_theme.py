"""
XLSX Theme & Styling Constants.

Centralizes ALL visual constants for the XLSX Comprehensive Quality Report.
Every other xlsx_* module imports from here -- no hardcoded colors/fonts elsewhere.

Spec reference: REQUIREMENTS-xlsx-report.md Section 六
"""

import platform
import math
from typing import Dict, Tuple, Optional

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =============================================================================
# 6.1  Font Configuration
# =============================================================================

def _detect_cjk_font() -> str:
    """Return the best available CJK font for the current platform.

    Uses platform detection with fallback chain.  Fonts are checked against
    matplotlib's font_manager so we only return names matplotlib can resolve.
    """
    import matplotlib.font_manager as fm

    system = platform.system()
    if system == "Darwin":
        candidates = ["Hiragino Sans GB", "Heiti TC", "Arial Unicode MS", "PingFang SC"]
    elif system == "Windows":
        candidates = ["Microsoft YaHei", "SimHei", "SimSun", "Arial Unicode MS"]
    else:
        candidates = ["WenQuanYi Micro Hei", "Noto Sans CJK SC", "Droid Sans Fallback"]

    available = {f.name for f in fm.fontManager.ttflist}
    for name in candidates:
        if name in available:
            return name
    # Last resort – DejaVu Sans ships with matplotlib but has no CJK glyphs.
    return "Arial"


def get_fonts(language: str = "en") -> Dict[str, str]:
    """Return dict of font names keyed by role: heading, body, number, mono."""
    if language == "zh":
        cjk = _detect_cjk_font()
        return {"heading": cjk, "body": cjk, "number": "Arial", "mono": "Consolas"}
    return {"heading": "Calibri", "body": "Calibri", "number": "Arial", "mono": "Consolas"}


# =============================================================================
# 6.2  Font Size Hierarchy (pt)
# =============================================================================

FONT_SIZES: Dict[str, int] = {
    "h1": 18,
    "h2": 14,
    "h3": 12,
    "th": 10,
    "td": 10,
    "kpi_value": 24,
    "kpi_label": 9,
    "footnote": 8,
    "insight_title": 11,
    "insight_body": 10,
}

# =============================================================================
# 6.3  Color Palette (6-digit hex, no # prefix unless noted)
# =============================================================================

# Primary palette
PRIMARY        = "1e3a5f"
PRIMARY_LIGHT  = "2d5a8e"
ACCENT         = "3b82f6"
TEXT_PRIMARY    = "1e293b"
TEXT_SECONDARY  = "64748b"
BACKGROUND     = "ffffff"
ZEBRA          = "f1f5f9"
BORDER         = "cbd5e1"
BORDER_LIGHT   = "e2e8f0"

# Semantic palette: key -> (bg_hex, text_hex)
SEMANTIC: Dict[str, Tuple[str, str]] = {
    "success": ("dcfce7", "166534"),
    "good":    ("dbeafe", "1e40af"),
    "warning": ("fef3c7", "92400e"),
    "danger":  ("fee2e2", "991b1b"),
    "neutral": ("f1f5f9", "475569"),
}

# Risk colors: key -> (bg_hex, text_hex)
RISK_COLORS: Dict[str, Tuple[str, str]] = {
    "Critical":  ("fee2e2", "991b1b"),
    "Warning":   ("fef3c7", "92400e"),
    "Attention": ("fef9c3", "854d0e"),
}

# Trend colors
TREND_POSITIVE = "166534"
TREND_NEGATIVE = "991b1b"
TREND_NEUTRAL  = "64748b"

# Chart colors for openpyxl native charts (no # prefix, 6-digit hex)
CHART_COLORS_HEX = [
    "3b82f6", "22c55e", "eab308", "ef4444",
    "8b5cf6", "06b6d4", "f97316", "ec4899",
    "14b8a6", "a855f7", "f43f5e", "84cc16",
]

# Native chart sizing (in cm for openpyxl)
CHART_WIDTH_CM  = 18        # standard chart width
CHART_HEIGHT_CM = 12        # standard chart height
CHART_SMALL_W   = 9         # small chart (gauge, sparkline)
CHART_SMALL_H   = 8         # small chart height

# Row computation for chart placement
DEFAULT_ROW_HEIGHT_PT = 15.0  # Excel default row height in points
CM_PER_PT = 0.0353            # centimeters per point

def rows_for_chart(height_cm: float) -> int:
    """Compute rows consumed by a chart of given height in cm."""
    row_height_cm = DEFAULT_ROW_HEIGHT_PT * CM_PER_PT
    return int(math.ceil(height_cm / row_height_cm)) + 2  # +2 padding

CHART_ROWS_STANDARD = rows_for_chart(CHART_HEIGHT_CM)  # ~25 for 12cm
CHART_ROWS_SMALL    = rows_for_chart(CHART_SMALL_H)    # ~17 for 8cm
CHART_ROWS_CONSUMED = CHART_ROWS_STANDARD              # backwards compat

# Tab colors for each sheet
TAB_COLORS = [
    "1e3a5f",  # Executive: dark blue
    "3b82f6",  # Incidents: blue
    "ef4444",  # SLA: red
    "22c55e",  # Changes: green
    "8b5cf6",  # Requests: purple
    "eab308",  # Problems: yellow
    "06b6d4",  # Cross-process: cyan
    "f97316",  # Personnel: orange
    "ec4899",  # Time: pink
    "14b8a6",  # Actions: teal
]

# AI Insight block colors
INSIGHT_BG     = "f0f9ff"
INSIGHT_BORDER = "3b82f6"
INSIGHT_TEXT   = "334155"

# KPI Card colors
KPI_CARD_BG     = "f8fafc"
KPI_CARD_BORDER = "e2e8f0"

# =============================================================================
# 6.4  Row Heights (pt)
# =============================================================================

ROW_HEIGHTS: Dict[str, float] = {
    "h1": 36,
    "h2": 22,
    "h3": 26,
    "th": 20,
    "td": 18,
    "spacer": 10,
    "insight": 20,
}

# =============================================================================
# 6.5  Column Widths (character units)
# =============================================================================

COL_WIDTHS: Dict[str, float] = {
    "ticket_id": 22,
    "short_text": 12,
    "percentage": 10,
    "count": 10,
    "time_value": 14,
    "long_text": 30,
    "person": 16,
}

# =============================================================================
# Pre-built openpyxl Style Objects
# =============================================================================

def _font(name: str, size: int, bold: bool = False, italic: bool = False,
          color: str = TEXT_PRIMARY) -> Font:
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


class XlsxStyles:
    """Pre-built openpyxl style objects for a given language. Create once per workbook."""

    def __init__(self, language: str = "en"):
        self.language = language
        fonts = get_fonts(language)
        hf = fonts["heading"]
        bf = fonts["body"]
        nf = fonts["number"]
        mf = fonts["mono"]

        # Fonts
        self.font_h1 = _font(hf, FONT_SIZES["h1"], bold=True, color=PRIMARY)
        self.font_h2 = _font(bf, FONT_SIZES["h2"], color=TEXT_SECONDARY)
        self.font_h3 = _font(hf, FONT_SIZES["h3"], bold=True, color=PRIMARY)
        self.font_th = _font(bf, FONT_SIZES["th"], bold=True, color="ffffff")
        self.font_td = _font(bf, FONT_SIZES["td"])
        self.font_td_num = _font(nf, FONT_SIZES["td"])
        self.font_td_mono = _font(mf, FONT_SIZES["td"])
        self.font_kpi_val = _font(nf, FONT_SIZES["kpi_value"], bold=True)
        self.font_kpi_label = _font(bf, FONT_SIZES["kpi_label"], color=TEXT_SECONDARY)
        self.font_footnote = _font(bf, FONT_SIZES["footnote"], italic=True, color=TEXT_SECONDARY)
        self.font_insight_title = _font(hf, FONT_SIZES["insight_title"], bold=True, color=PRIMARY)
        self.font_insight_body = _font(bf, FONT_SIZES["insight_body"], color=INSIGHT_TEXT)

        # Fills
        self.fill_header = _fill(PRIMARY)
        self.fill_zebra = _fill(ZEBRA)
        self.fill_white = _fill(BACKGROUND)
        self.fill_insight = _fill(INSIGHT_BG)
        self.fill_kpi_card = _fill(KPI_CARD_BG)

        # Alignment
        self.align_left = Alignment(horizontal="left", vertical="center")
        self.align_center = Alignment(horizontal="center", vertical="center")
        self.align_right = Alignment(horizontal="right", vertical="center")
        self.align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.align_h1 = Alignment(horizontal="left", vertical="center")

        # Borders
        self.border_header = Border(bottom=Side(style="thin", color=BORDER))
        self.border_row = Border(bottom=Side(style="thin", color=BORDER_LIGHT))
        self.border_kpi = Border(
            top=Side(style="thin", color=KPI_CARD_BORDER),
            bottom=Side(style="thin", color=KPI_CARD_BORDER),
            left=Side(style="thin", color=KPI_CARD_BORDER),
            right=Side(style="thin", color=KPI_CARD_BORDER),
        )

    # -- Semantic helpers -----------------------------------------------------

    def semantic_fill(self, level: str) -> PatternFill:
        bg, _ = SEMANTIC.get(level, SEMANTIC["neutral"])
        return _fill(bg)

    def semantic_font(self, level: str, size: int = FONT_SIZES["td"],
                      bold: bool = False) -> Font:
        _, fg = SEMANTIC.get(level, SEMANTIC["neutral"])
        return _font("Arial", size, bold=bold, color=fg)

    def risk_fill(self, priority: str) -> PatternFill:
        bg, _ = RISK_COLORS.get(priority, ("f1f5f9", "475569"))
        return _fill(bg)

    def risk_font(self, priority: str) -> Font:
        _, fg = RISK_COLORS.get(priority, ("f1f5f9", "475569"))
        return _font("Arial", FONT_SIZES["td"], bold=True, color=fg)

    def trend_font(self, direction: str, positive_means: str = "up") -> Font:
        """Return font for trend arrows.

        Args:
            direction: arrow character "↑", "↓", or "→"
            positive_means: 'up' = increase is good (SLA), 'down' = decrease is good (incidents)
        """
        if direction == "→":
            color = TREND_NEUTRAL
        elif direction == "↑":
            color = TREND_POSITIVE if positive_means == "up" else TREND_NEGATIVE
        elif direction == "↓":
            color = TREND_POSITIVE if positive_means == "down" else TREND_NEGATIVE
        else:
            color = TREND_NEUTRAL
        return _font("Arial", FONT_SIZES["td"], bold=True, color=color)


# =============================================================================
# SLA Rating Helpers
# =============================================================================

def sla_level(rate: float) -> str:
    """Return semantic level for an SLA rate (0.0 - 1.0 scale)."""
    if rate >= 0.95:
        return "success"
    if rate >= 0.90:
        return "good"
    if rate >= 0.80:
        return "warning"
    return "danger"


def rating_text(rate: float, language: str = "en") -> str:
    """Return rating label for an SLA rate (0.0 - 1.0 scale)."""
    labels = {
        "en": {0.95: "Excellent", 0.90: "Good", 0.80: "Fair", 0: "At Risk"},
        "zh": {0.95: "优秀", 0.90: "良好", 0.80: "一般", 0: "风险"},
    }
    lang_labels = labels.get(language, labels["en"])
    for threshold in sorted(lang_labels.keys(), reverse=True):
        if rate >= threshold:
            return lang_labels[threshold]
    return lang_labels[0]


def efficiency_level(rating: str) -> str:
    """Map Chinese/English rating text to semantic level."""
    mapping = {
        "优秀": "success", "Excellent": "success",
        "良好": "good", "Good": "good",
        "一般": "neutral", "Fair": "neutral",
        "需改进": "danger", "Needs Improvement": "danger",
        "风险": "danger", "At Risk": "danger",
    }
    return mapping.get(rating, "neutral")


# =============================================================================
# Data Formatting Helpers
# =============================================================================

def format_duration(minutes: float, language: str = "en") -> str:
    """Format minutes into human-readable duration."""
    if minutes is None or minutes != minutes:  # NaN check
        return "N/A"
    if minutes < 0:
        return "N/A"
    if language == "zh":
        if minutes < 60:
            return f"{minutes:.1f}分钟"
        hours = minutes / 60
        if hours < 48:
            return f"{hours:.1f}小时"
        return f"{hours / 24:.1f}天"
    else:
        if minutes < 60:
            return f"{minutes:.1f}min"
        hours = minutes / 60
        if hours < 48:
            return f"{hours:.1f}h"
        return f"{hours / 24:.1f}d"


def format_pct(value: float) -> str:
    """Format a 0.0-1.0 float as percentage string."""
    if value is None or value != value:
        return "N/A"
    return f"{value:.1%}"


def format_number(value) -> str:
    """Format integer with comma separator."""
    if value is None:
        return "N/A"
    return f"{int(value):,}"

